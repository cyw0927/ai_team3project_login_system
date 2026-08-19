from .common import *

@admin_required
def admin_teams(request):
    selected_round_id = request.GET.get("round", "").strip()
    query = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "all").strip() or "all"
    rounds = EvaluationRound.objects.all().order_by("-start_at")
    selected_round = None
    if selected_round_id:
        selected_round = rounds.filter(pk=selected_round_id).first()
    if selected_round is None:
        selected_round = _current_round()

    base_teams_qs = Team.objects.filter(evaluation_round=selected_round) if selected_round else Team.objects.none()
    all_round_teams = list(base_teams_qs.order_by("name"))

    teams_qs = base_teams_qs
    if query:
        teams_qs = teams_qs.filter(
            Q(name__icontains=query)
            | Q(project_title__icontains=query)
            | Q(memberships__student__user__first_name__icontains=query)
            | Q(memberships__student__user__email__icontains=query)
        ).distinct()
    if status_filter == "active":
        teams_qs = teams_qs.filter(is_active=True)
    elif status_filter == "inactive":
        teams_qs = teams_qs.filter(is_active=False)
    teams = list(teams_qs.order_by("name"))

    # 화면에 표시할 팀원은 검색 결과에 관계없이 해당 팀 전체 구성원을 보여준다.
    display_team_ids = [team.id for team in teams]
    memberships = (
        TeamMembership.objects.filter(team_id__in=display_team_ids)
        .select_related("student__user", "team")
        .order_by("team__name", "student__user__first_name", "student__user__username")
    )
    member_map = {team.id: [] for team in teams}
    for membership in memberships:
        member_map.setdefault(membership.team_id, []).append(membership.student)
    for team in teams:
        team.members = member_map.get(team.id, [])
        team.member_count = len(team.members)
        team.status_display = "활성" if team.is_active else "비활성"
        team.leader = team.members[0] if team.members else None

    active_students = Student.objects.filter(is_active=True, user__is_active=True)
    assigned_ids = set()
    all_membership_counts = {}
    if selected_round:
        round_memberships = TeamMembership.objects.filter(team__evaluation_round=selected_round)
        assigned_ids = set(round_memberships.values_list("student_id", flat=True))
        for row in round_memberships.values("team_id").annotate(total=Count("id")):
            all_membership_counts[row["team_id"]] = row["total"]

    total_team_count = len(all_round_teams)
    assigned_count = len(assigned_ids)
    average_members = round(assigned_count / total_team_count, 1) if total_team_count else 0
    unassigned_students = list(
        active_students.exclude(id__in=assigned_ids)
        .select_related("user")
        .order_by("user__first_name", "user__username")[:8]
    )
    stats = {
        "team_count": total_team_count,
        "assigned_count": assigned_count,
        "active_student_count": active_students.count(),
        "average_members": average_members,
        "unassigned_count": active_students.exclude(id__in=assigned_ids).count(),
        "active_team_count": sum(1 for team in all_round_teams if team.is_active),
    }
    return render(
        request,
        "admin_ui/teams.html",
        _base_context(
            teams=teams,
            rounds=rounds,
            selected_round=selected_round,
            query=query,
            status_filter=status_filter,
            unassigned_students=unassigned_students,
            stats=stats,
        ),
    )

@admin_required
def admin_teams_excel_export(request):
    """선택한 회차의 팀 구성표를 Excel(.xlsx)로 내려준다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    round_id = request.GET.get("round", "").strip()
    evaluation_round = EvaluationRound.objects.filter(pk=round_id).first() if round_id else _current_round()
    if not evaluation_round:
        messages.error(request, "내보낼 평가 회차가 없습니다.")
        return redirect("admin_teams")

    teams = list(Team.objects.filter(evaluation_round=evaluation_round).order_by("name"))
    memberships = (
        TeamMembership.objects.filter(team__in=teams)
        .select_related("team", "student__user")
        .order_by("team__name", "student__user__first_name", "student__user__username")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "팀 구성"
    headers = ["회차", "팀명", "프로젝트", "상태", "학생명", "이메일", "소속"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    member_map = {team.id: [] for team in teams}
    for membership in memberships:
        member_map.setdefault(membership.team_id, []).append(membership.student)

    for team in teams:
        members = member_map.get(team.id, [])
        if not members:
            ws.append([evaluation_round.name, team.name, team.project_title or "", "활성" if team.is_active else "비활성", "", "", ""])
            continue
        for student in members:
            ws.append([
                evaluation_round.name,
                team.name,
                team.project_title or "",
                "활성" if team.is_active else "비활성",
                student.user.get_full_name() or student.user.username,
                student.user.email,
                student.affiliation or "",
            ])

    widths = {"A": 22, "B": 14, "C": 30, "D": 12, "E": 16, "F": 30, "G": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="teams_round_{evaluation_round.id}.xlsx"'
    return response

@admin_required
@require_POST
@transaction.atomic
def admin_team_create(request):
    if request.method != "POST":
        return _redirect_back(request, "admin_teams")
    form = TeamForm(request.POST, rounds=EvaluationRound.objects.all())
    if not form.is_valid():
        for errors in form.errors.values():
            for error in errors:
                messages.error(request, error)
        return _redirect_back(request, "admin_teams")
    team = form.save()
    messages.success(request, f"{team.name}을(를) 생성했습니다.")
    return _redirect_back(request, f"/management/teams/?round={team.evaluation_round_id}")

@admin_required
@require_POST
@transaction.atomic
def admin_team_update(request, team_id):
    if request.method != "POST":
        return _redirect_back(request, "admin_teams")
    team = get_object_or_404(Team, pk=team_id)
    form = TeamForm(request.POST, instance=team, rounds=EvaluationRound.objects.all())
    if not form.is_valid():
        for errors in form.errors.values():
            for error in errors:
                messages.error(request, error)
        return _redirect_back(request, f"/management/teams/?round={team.evaluation_round_id}")
    team = form.save()
    messages.success(request, f"{team.name} 정보를 수정했습니다.")
    return _redirect_back(request, f"/management/teams/?round={team.evaluation_round_id}")

@admin_required
@require_POST
@transaction.atomic
def admin_team_delete(request, team_id):
    if request.method != "POST":
        return _redirect_back(request, "admin_teams")
    team = get_object_or_404(Team, pk=team_id)
    round_id = team.evaluation_round_id
    name = team.name
    if team.memberships.exists():
        messages.error(request, "팀원이 배정된 팀은 바로 삭제할 수 없습니다. 먼저 팀원을 이동하거나 배정을 해제하세요.")
        return _redirect_back(request, f"/management/teams/?round={round_id}")
    team.delete()
    messages.success(request, f"{name}을(를) 삭제했습니다.")
    return _redirect_back(request, f"/management/teams/?round={round_id}")

@admin_required
def admin_team_assignment(request):
    rounds = EvaluationRound.objects.all().order_by("-start_at")
    selected_round_id = request.GET.get("round") or request.POST.get("round_id")
    selected_round = rounds.filter(pk=selected_round_id).first() if selected_round_id else _current_round()

    teams = list(Team.objects.filter(evaluation_round=selected_round, is_active=True).order_by("name")) if selected_round else []
    memberships = (
        TeamMembership.objects.filter(team__in=teams)
        .select_related("student__user", "team")
        if teams else TeamMembership.objects.none()
    )
    member_map = {team.id: [] for team in teams}
    assigned_ids = set()
    for membership in memberships:
        member_map.setdefault(membership.team_id, []).append(membership)
        assigned_ids.add(membership.student_id)
    for team in teams:
        team.membership_rows = member_map.get(team.id, [])
        team.members = [membership.student for membership in team.membership_rows]

    unassigned_students = list(
        Student.objects.filter(is_active=True, user__is_active=True)
        .exclude(id__in=assigned_ids)
        .select_related("user")
        .order_by("user__first_name", "user__username")
    )

    preview_teams = request.session.get("team_assignment_preview", [])
    if preview_teams and str(request.session.get("team_assignment_round_id")) != str(getattr(selected_round, "id", "")):
        preview_teams = []

    pot_quality = None
    pot_quality_warnings = []
    if preview_teams and preview_teams[0].get("assignment_rule") == "pot":
        unseeded_count = sum(team.get("pot_counts", {}).get("U", 0) for team in preview_teams)
        if unseeded_count:
            pot_quality_warnings.append(f"Seed 미분류 {unseeded_count}명은 U로 균등 랜덤 배치")
        for grade in ("A", "B", "C", "D"):
            counts = [team.get("pot_counts", {}).get(grade, 0) for team in preview_teams]
            spread = max(counts) - min(counts) if counts else 0
            if spread > 1:
                pot_quality_warnings.append(
                    f"{grade} POT 팀간 차이 {spread}명"
                )
        if pot_quality_warnings:
            pot_quality = {
                "status": "warning",
                "label": "포트 분산 확인 필요",
                "message": "일부 포트가 특정 팀에 상대적으로 몰렸습니다. 다시 추첨하면 다른 조합을 확인할 수 있습니다.",
            }
        else:
            pot_quality = {
                "status": "good",
                "label": "포트 균형 양호",
                "message": "A/B/C/D 포트가 팀별로 가능한 범위 안에서 고르게 분산되었습니다.",
            }

    total_students = Student.objects.filter(is_active=True, user__is_active=True).count()
    assigned_count = len(assigned_ids)
    unassigned_count = len(unassigned_students)
    team_count = len(teams)
    average_team_size = round(assigned_count / team_count, 1) if team_count else 0
    largest_team_size = max((len(team.members) for team in teams), default=0)
    smallest_team_size = min((len(team.members) for team in teams), default=0) if teams else 0

    auto_settings = request.session.get("team_assignment_auto_settings", {})
    if str(auto_settings.get("round_id", "")) != str(getattr(selected_round, "id", "")):
        auto_settings = {}

    # 현재 탭을 서버에서도 기억해 새로고침/POST 후에도 같은 탭을 유지한다.
    active_tab = request.GET.get("tab", "").strip().lower()
    if active_tab not in {"manual", "auto"}:
        active_tab = "auto" if preview_teams else "manual"

    return render(
        request,
        "admin_ui/team_assignment.html",
        _base_context(
            rounds=rounds,
            selected_round=selected_round,
            unassigned_students=unassigned_students,
            teams=teams,
            preview_teams=preview_teams,
            pot_quality=pot_quality,
            pot_quality_warnings=pot_quality_warnings,
            form_data={
                "team_count": auto_settings.get("team_count", len(teams) or ""),
                "assignment_rule": auto_settings.get("assignment_rule", "seed"),
                "avoid_previous": auto_settings.get("avoid_previous", "1"),
            },
            total_students=total_students,
            assigned_count=assigned_count,
            unassigned_count=unassigned_count,
            team_count=team_count,
            average_team_size=average_team_size,
            largest_team_size=largest_team_size,
            smallest_team_size=smallest_team_size,
            active_tab=active_tab,
        ),
    )

@admin_required
@require_POST
@transaction.atomic
def admin_manual_assign(request):
    if request.method != "POST":
        return _redirect_back(request, "admin_team_assignment")
    round_id = request.POST.get("round_id")
    team_id = request.POST.get("team_id")
    student_ids = request.POST.getlist("student_ids")
    evaluation_round = get_object_or_404(EvaluationRound, pk=round_id)
    team = get_object_or_404(Team, pk=team_id, evaluation_round=evaluation_round)
    if not student_ids:
        messages.error(request, "배정할 수강생을 선택하세요.")
        return _redirect_back(request, f"/management/team-assignment/?round={round_id}")
    students = Student.objects.filter(pk__in=student_ids, is_active=True, user__is_active=True)
    for student in students:
        TeamMembership.objects.filter(student=student, team__evaluation_round=evaluation_round).delete()
        TeamMembership.objects.create(team=team, student=student)
    messages.success(request, f"{students.count()}명을 {team.name}에 배정했습니다.")
    return _redirect_back(request, f"/management/team-assignment/?round={round_id}")

@admin_required
@require_POST
@transaction.atomic
def admin_manual_unassign(request):
    if request.method != "POST":
        return _redirect_back(request, "admin_team_assignment")
    round_id = request.POST.get("round_id")
    membership_ids = request.POST.getlist("membership_ids")
    if not membership_ids:
        messages.error(request, "배정을 해제할 팀원을 선택하세요.")
        return _redirect_back(request, f"/management/team-assignment/?round={round_id}")
    deleted, _ = TeamMembership.objects.filter(
        pk__in=membership_ids,
        team__evaluation_round_id=round_id,
    ).delete()
    messages.success(request, "선택한 팀원의 배정을 해제했습니다.")
    return _redirect_back(request, f"/management/team-assignment/?round={round_id}")

@admin_required
@require_POST
@transaction.atomic
def admin_team_dissolve_all(request):
    """선택 회차의 현재 팀 배정을 한 번에 해체한다.

    평가 데이터와 팀 식별자는 보존해야 하므로 Team 자체를 삭제하지 않고
    TeamMembership만 제거한다. 이후 수동/자동 편성을 즉시 다시 할 수 있다.
    """
    round_id = request.POST.get("round_id", "").strip()
    evaluation_round = get_object_or_404(EvaluationRound, pk=round_id)

    memberships = TeamMembership.objects.filter(team__evaluation_round=evaluation_round)
    member_count = memberships.count()
    team_count = Team.objects.filter(evaluation_round=evaluation_round, is_active=True).count()

    if member_count == 0:
        messages.info(request, "현재 해체할 팀 배정이 없습니다.")
        return _redirect_back(
            request,
            f"/management/team-assignment/?round={evaluation_round.id}&tab=manual",
        )

    memberships.delete()

    # 해당 회차의 자동 편성 미리보기도 폐기해 오래된 상태가 다시 확정되지 않게 한다.
    if str(request.session.get("team_assignment_round_id", "")) == str(evaluation_round.id):
        request.session.pop("team_assignment_preview", None)
        request.session.pop("team_assignment_round_id", None)
        request.session.modified = True

    messages.success(
        request,
        f"{team_count}개 팀의 {member_count}명 배정을 모두 해체했습니다. 팀 이름은 재편성을 위해 유지됩니다.",
    )
    return _redirect_back(
        request,
        f"/management/team-assignment/?round={evaluation_round.id}&tab=manual",
    )


@admin_required
@require_POST
def admin_team_member_role_update(request, membership_id):
    membership = get_object_or_404(
        TeamMembership.objects.select_related("team__evaluation_round", "student__user"),
        pk=membership_id,
    )
    role = request.POST.get("role", "").strip()
    if len(role) > 100:
        messages.error(request, "담당 역할은 100자 이하로 입력해 주세요.")
        return _redirect_back(request, f"/management/team-assignment/?round={membership.team.evaluation_round_id}&tab=manual")
    membership.role = role
    membership.save(update_fields=["role", "updated_at"])
    messages.success(request, f"{membership.student.name}님의 담당 역할을 저장했습니다.")
    return _redirect_back(request, f"/management/team-assignment/?round={membership.team.evaluation_round_id}&tab=manual#manual")

@admin_required
@require_POST
def admin_auto_preview(request):
    if request.method != "POST":
        return _redirect_back(request, "admin_team_assignment")
    round_id = request.POST.get("round_id")
    evaluation_round = get_object_or_404(EvaluationRound, pk=round_id)
    try:
        team_count = int(request.POST.get("team_count", "0"))
    except ValueError:
        team_count = 0
    if team_count < 1:
        messages.error(request, "팀 수는 1개 이상이어야 합니다.")
        return _redirect_back(request, f"/management/team-assignment/?round={round_id}&tab=auto")

    students = list(Student.objects.filter(is_active=True, user__is_active=True).select_related("user"))
    if not students:
        messages.error(request, "편성할 활성 수강생이 없습니다.")
        return _redirect_back(request, f"/management/team-assignment/?round={round_id}&tab=auto")
    if team_count > len(students):
        messages.error(request, "팀 수는 활성 수강생 수보다 많을 수 없습니다.")
        return _redirect_back(request, f"/management/team-assignment/?round={round_id}&tab=auto")

    previous_team_map = {}
    if request.POST.get("avoid_previous") == "1":
        previous_round = _previous_round_for(evaluation_round)
        if previous_round:
            previous_team_map = dict(
                TeamMembership.objects.filter(team__evaluation_round=previous_round)
                .values_list("student_id", "team_id")
            )

    assignment_rule = request.POST.get("assignment_rule", "seed")
    avoid_previous = request.POST.get("avoid_previous", "1")
    request.session["team_assignment_auto_settings"] = {
        "round_id": evaluation_round.id,
        "team_count": team_count,
        "assignment_rule": assignment_rule,
        "avoid_previous": avoid_previous,
    }
    request.session.modified = True
    seed_scores = {}
    previous_round = _previous_round_for(evaluation_round)
    if assignment_rule in {"seed", "pot"} and previous_round:
        # 이전 회차 중 실제 원본 평가가 있는 회차는 최신 결과로 다시 집계한다.
        previous_rounds = EvaluationRound.objects.filter(
            start_at__lt=evaluation_round.start_at,
            status=EvaluationRound.Status.ENDED,
        ).order_by("start_at")
        for prior_round in previous_rounds:
            has_raw_scores = (
                TeamEvaluationScore.objects.filter(evaluation__evaluation_round=prior_round).exists()
                or PersonalEvaluationScore.objects.filter(evaluation__evaluation_round=prior_round).exists()
            )
            if has_raw_scores:
                _recalculate_round_results(prior_round)

        # RFP의 누적 시드 요구에 맞춰 이전 모든 회차 final_score의 학생별 평균을 사용한다.
        seed_scores = _cumulative_seed_scores_before(evaluation_round)

    grade_map = {}
    pot_counts = {}
    if assignment_rule == "seed" and seed_scores:
        buckets = _snake_seed_assignment(students, team_count, seed_scores)
        messages.info(request, "이전 평가들의 개인 최종점수 누적 평균을 기준으로 Z식 시드 편성을 적용했습니다.")
    elif assignment_rule == "pot" and seed_scores:
        buckets, grade_map, pot_counts = _pot_seed_assignment(
            students,
            team_count,
            seed_scores,
            previous_team_map,
        )
        messages.info(
            request,
            "누적 Seed 순위를 A/B/C/D 포트로 나눈 뒤 포트별 랜덤 추첨 방식으로 편성했습니다.",
        )
    else:
        if assignment_rule in {"seed", "pot"}:
            messages.warning(request, "이전 회차 성적 데이터가 없어 균등 랜덤 방식으로 미리보기를 만들었습니다.")
        buckets = _balanced_random_assignment(students, team_count, previous_team_map)
    seed_rank_map = {}

    if seed_scores:
        ranked_students = sorted(
            [student for student in students if student.id in seed_scores],
            key=lambda student: (float(seed_scores[student.id]), -student.id),
            reverse=True,
        )
        seed_rank_map = {
            student.id: rank
            for rank, student in enumerate(ranked_students, start=1)
        }

    seed_number_map = {
        student_id: ((rank - 1) // team_count) + 1
        for student_id, rank in seed_rank_map.items()
    }

    preview = []
    for idx, bucket in enumerate(buckets, start=1):
        members = [
            {
                "name": s.name,
                "seed_rank": seed_rank_map.get(s.id),
                "seed_number": seed_number_map.get(s.id),
                "seed_score": float(seed_scores[s.id]) if s.id in seed_scores else None,
                "pot_grade": grade_map.get(s.id),
            }
            for s in bucket
        ]
        team_pot_counts = {"A": 0, "B": 0, "C": 0, "D": 0, "U": 0}
        if assignment_rule == "pot":
            for member in members:
                grade = member.get("pot_grade")
                if grade in team_pot_counts:
                    team_pot_counts[grade] += 1

        preview.append({
            "name": f"{idx}팀",
            "student_ids": [s.id for s in bucket],
            "members": members,
            "pot_counts": team_pot_counts,
            "assignment_rule": assignment_rule,
        })
    request.session["team_assignment_preview"] = preview
    request.session["team_assignment_round_id"] = evaluation_round.id
    request.session.modified = True
    messages.success(request, "자동 편성 미리보기를 만들었습니다. 확정 전까지 DB에는 반영되지 않습니다.")
    return _redirect_back(request, f"/management/team-assignment/?round={round_id}&tab=auto#auto")

@admin_required
@require_POST
@transaction.atomic
def admin_auto_confirm(request):
    if request.method != "POST":
        return _redirect_back(request, "admin_team_assignment")
    round_id = request.POST.get("round_id")
    evaluation_round = get_object_or_404(EvaluationRound, pk=round_id)
    preview = request.session.get("team_assignment_preview", [])
    preview_round_id = request.session.get("team_assignment_round_id")
    if not preview or str(preview_round_id) != str(evaluation_round.id):
        messages.error(request, "확정할 자동 편성 미리보기가 없습니다.")
        return _redirect_back(request, f"/management/team-assignment/?round={round_id}&tab=auto")

    TeamMembership.objects.filter(team__evaluation_round=evaluation_round).delete()
    existing = {team.name: team for team in Team.objects.filter(evaluation_round=evaluation_round)}
    used_team_ids = []
    for team_data in preview:
        team = existing.get(team_data["name"])
        if team is None:
            team = Team.objects.create(
                evaluation_round=evaluation_round,
                name=team_data["name"],
                is_active=True,
            )
        else:
            if not team.is_active:
                team.is_active = True
                team.save(update_fields=["is_active", "updated_at"])
        used_team_ids.append(team.id)
        students = Student.objects.filter(pk__in=team_data["student_ids"], is_active=True, user__is_active=True)
        TeamMembership.objects.bulk_create([TeamMembership(team=team, student=student) for student in students])

    Team.objects.filter(evaluation_round=evaluation_round).exclude(pk__in=used_team_ids).update(is_active=False)
    request.session.pop("team_assignment_preview", None)
    request.session.pop("team_assignment_round_id", None)
    request.session.modified = True
    messages.success(request, "자동 팀 편성을 확정하여 DB에 반영했습니다.")
    return _redirect_back(request, f"/management/team-assignment/?round={round_id}&tab=auto")
