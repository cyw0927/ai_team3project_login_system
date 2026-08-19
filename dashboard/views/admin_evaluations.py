from .common import *


def _template_url(template):
    """수정/추가 후 현재 평가 유형 탭과 선택 템플릿을 그대로 유지한다."""
    return (
        f"/management/evaluation-templates/"
        f"?type={template.evaluation_type}&template={template.id}"
    )


def _show_form_errors(request, form):
    for errors in form.errors.values():
        for error in errors:
            messages.error(request, error)


def _submission_progress(evaluation_round, memberships_qs, active_teams, attendance_map):
    """Count only valid evaluator-target pairs for the selected round.

    Old/draft data can contain evaluations whose evaluator is no longer a member or
    whose target is inactive. Counting raw rows makes the numerator exceed the
    required denominator, so progress is derived from the same eligibility rules
    used to build required evaluation pairs.
    """
    memberships = list(memberships_qs)
    members_by_team = {}
    for membership in memberships:
        members_by_team.setdefault(membership.team_id, []).append(membership.student_id)

    submitted_team_pairs = set(
        TeamEvaluation.objects.filter(
            evaluation_round=evaluation_round,
            is_submitted=True,
        ).values_list("evaluator_id", "target_team_id")
    )
    submitted_personal_pairs = set(
        PersonalEvaluation.objects.filter(
            evaluation_round=evaluation_round,
            is_submitted=True,
        ).values_list("evaluator_id", "target_student_id")
    )

    team_required = team_completed = 0
    personal_required = personal_completed = 0

    for membership in memberships:
        evaluator_id = membership.student_id
        team_id = membership.team_id
        attendance_status = attendance_map.get(evaluator_id, RoundAttendance.Status.PRESENT)

        if attendance_status == RoundAttendance.Status.PRESENT:
            for target_team in active_teams:
                if target_team.id == team_id:
                    continue
                team_required += 1
                team_completed += (evaluator_id, target_team.id) in submitted_team_pairs

        for target_student_id in members_by_team.get(team_id, []):
            if target_student_id == evaluator_id:
                continue
            personal_required += 1
            personal_completed += (evaluator_id, target_student_id) in submitted_personal_pairs

    return {
        "team_required": team_required,
        "team_completed": team_completed,
        "personal_required": personal_required,
        "personal_completed": personal_completed,
    }


def _evaluation_review_flags(evaluation_round):
    """Return submitted evaluations whose score pattern deserves a quick admin review.

    These are hints, not misconduct judgments. Only clear patterns are surfaced so the
    result page stays useful instead of noisy.
    """
    flags = []

    def add_flag(kind, evaluator, target, scores):
        if len(scores) < 2:
            return
        avg = sum(scores) / len(scores)
        unique = set(scores)
        reason = None
        if len(unique) == 1:
            reason = f"모든 항목 {scores[0]}점"
        elif avg <= 1.5:
            reason = "낮은 점수 집중"
        elif avg >= 4.8:
            reason = "높은 점수 집중"
        if reason:
            flags.append({
                "kind": kind,
                "evaluator": evaluator.name,
                "target": target,
                "average": round(avg, 2),
                "reason": reason,
            })

    team_evaluations = (
        TeamEvaluation.objects.filter(evaluation_round=evaluation_round, is_submitted=True)
        .select_related("evaluator__user", "target_team")
        .prefetch_related("scores")
    )
    for evaluation in team_evaluations:
        add_flag(
            "팀",
            evaluation.evaluator,
            evaluation.target_team.name,
            [item.score for item in evaluation.scores.all()],
        )

    personal_evaluations = (
        PersonalEvaluation.objects.filter(evaluation_round=evaluation_round, is_submitted=True)
        .select_related("evaluator__user", "target_student__user")
        .prefetch_related("scores")
    )
    for evaluation in personal_evaluations:
        add_flag(
            "개인",
            evaluation.evaluator,
            evaluation.target_student.name,
            [item.score for item in evaluation.scores.all()],
        )

    return flags

@admin_required
def admin_evaluation_templates(request):
    selected_template_id = request.GET.get("template", "").strip()
    type_filter = request.GET.get("type", "team").strip()
    if type_filter not in {EvaluationTemplate.EvaluationType.TEAM, EvaluationTemplate.EvaluationType.PERSONAL}:
        type_filter = EvaluationTemplate.EvaluationType.TEAM

    # 회차 적용용 스냅샷은 내부 보존용이므로 템플릿 라이브러리에는 중복 노출하지 않는다.
    templates = list(
        EvaluationTemplate.objects.filter(evaluation_round__isnull=True)
        .select_related("evaluation_round")
        .prefetch_related("criteria")
        .order_by("evaluation_type", "name")
    )
    for template in templates:
        template.type_display = template.get_evaluation_type_display()
        template.status_display = "사용 중" if template.is_active else "비활성"
        template.criterion_count = template.criteria.count()

    filtered_templates = [t for t in templates if t.evaluation_type == type_filter]
    selected_template = None
    if selected_template_id:
        selected_template = next((t for t in filtered_templates if str(t.id) == selected_template_id), None)
    if selected_template is None and filtered_templates:
        selected_template = filtered_templates[0]

    criteria = []
    if selected_template:
        criteria = list(selected_template.criteria.all().order_by("order", "id"))

    rounds = EvaluationRound.objects.all().order_by("-start_at")
    stats = {
        "total": len(templates),
        "team": sum(t.evaluation_type == EvaluationTemplate.EvaluationType.TEAM for t in templates),
        "personal": sum(t.evaluation_type == EvaluationTemplate.EvaluationType.PERSONAL for t in templates),
        "active": sum(t.is_active for t in templates),
    }
    return render(
        request,
        "admin_ui/evaluation_templates.html",
        _base_context(
            evaluation_templates=templates,
            filtered_templates=filtered_templates,
            selected_template=selected_template,
            criteria=criteria,
            rounds=rounds,
            stats=stats,
            type_filter=type_filter,
            required_count=sum(c.is_required for c in criteria),
            optional_count=sum(not c.is_required for c in criteria),
        ),
    )

@admin_required
@require_POST
@transaction.atomic
def admin_evaluation_template_create(request):
    form = EvaluationTemplateForm(request.POST, rounds=EvaluationRound.objects.all())
    if not form.is_valid():
        _show_form_errors(request, form)
        return _redirect_back(request, "admin_evaluation_templates")
    template = form.save()
    messages.success(request, f"{template.name} 템플릿을 생성했습니다.")
    return redirect(_template_url(template))

@admin_required
@require_POST
@transaction.atomic
def admin_evaluation_template_update(request, template_id):
    template = get_object_or_404(EvaluationTemplate, pk=template_id)
    form = EvaluationTemplateForm(
        request.POST,
        instance=template,
        rounds=EvaluationRound.objects.all(),
    )
    if not form.is_valid():
        _show_form_errors(request, form)
        return redirect(_template_url(template))
    template = form.save()
    messages.success(request, f"{template.name} 템플릿을 수정했습니다.")
    return redirect(_template_url(template))

@admin_required
@require_POST
@transaction.atomic
def admin_evaluation_template_toggle(request, template_id):
    template = get_object_or_404(EvaluationTemplate, pk=template_id)
    template.is_active = not template.is_active
    template.save(update_fields=["is_active", "updated_at"])
    messages.success(request, f"{template.name} 템플릿 상태를 변경했습니다.")
    return redirect(_template_url(template))

@admin_required
@require_POST
@transaction.atomic
def admin_evaluation_template_delete(request, template_id):
    template = get_object_or_404(EvaluationTemplate, pk=template_id)
    if template.criteria.filter(team_scores__isnull=False).exists() or template.criteria.filter(personal_scores__isnull=False).exists():
        messages.error(request, "이미 평가에 사용된 템플릿은 삭제할 수 없습니다. 비활성화하세요.")
        return redirect(_template_url(template))
    name = template.name
    template.delete()
    messages.success(request, f"{name} 템플릿을 삭제했습니다.")
    return _redirect_back(request, "admin_evaluation_templates")

@admin_required
@require_POST
@transaction.atomic
def admin_evaluation_criterion_create(request, template_id):
    template = get_object_or_404(EvaluationTemplate, pk=template_id)
    form = EvaluationCriterionForm(request.POST)
    if not form.is_valid():
        _show_form_errors(request, form)
        return redirect(_template_url(template))
    criterion = form.save(commit=False)
    criterion.template = template
    if not request.POST.get("order"):
        last = template.criteria.order_by("-order", "-id").first()
        criterion.order = (last.order + 1) if last else 1
    criterion.save()
    messages.success(request, f"{criterion.title} 문항을 추가했습니다.")
    return redirect(_template_url(template))

@admin_required
@require_POST
@transaction.atomic
def admin_evaluation_criterion_update(request, criterion_id):
    criterion = get_object_or_404(EvaluationCriterion.objects.select_related("template"), pk=criterion_id)
    form = EvaluationCriterionForm(request.POST, instance=criterion)
    if not form.is_valid():
        _show_form_errors(request, form)
        return redirect(_template_url(criterion.template))
    criterion = form.save()
    messages.success(request, f"{criterion.title} 문항을 수정했습니다.")
    return redirect(_template_url(criterion.template))

@admin_required
@require_POST
@transaction.atomic
def admin_evaluation_criterion_delete(request, criterion_id):
    criterion = get_object_or_404(EvaluationCriterion.objects.select_related("template"), pk=criterion_id)
    template_id = criterion.template_id
    if criterion.team_scores.exists() or criterion.personal_scores.exists():
        messages.error(request, "이미 평가에 사용된 문항은 삭제할 수 없습니다.")
        return redirect(_template_url(criterion.template))
    title = criterion.title
    template = criterion.template
    criterion.delete()
    messages.success(request, f"{title} 문항을 삭제했습니다.")
    return redirect(_template_url(template))

@admin_required
@require_POST
@transaction.atomic
def admin_evaluation_criterion_move(request, criterion_id, direction):
    criterion = get_object_or_404(EvaluationCriterion.objects.select_related("template"), pk=criterion_id)
    siblings = list(criterion.template.criteria.all().order_by("order", "id"))
    try:
        index = next(i for i, item in enumerate(siblings) if item.id == criterion.id)
    except StopIteration:
        return redirect(_template_url(criterion.template))
    target_index = index - 1 if direction == "up" else index + 1 if direction == "down" else index
    if 0 <= target_index < len(siblings) and target_index != index:
        target = siblings[target_index]
        criterion.order, target.order = target.order, criterion.order
        criterion.save(update_fields=["order", "updated_at"])
        target.save(update_fields=["order", "updated_at"])
    return redirect(_template_url(criterion.template))

@admin_required
@require_POST
@transaction.atomic
def admin_evaluation_criteria_reorder(request, template_id):
    template = get_object_or_404(EvaluationTemplate, pk=template_id)
    raw_ids = request.POST.get("ordered_ids", "")
    try:
        ordered_ids = [int(value) for value in raw_ids.split(",") if value.strip()]
    except ValueError:
        messages.error(request, "문항 순서 정보가 올바르지 않습니다.")
        return redirect(f"/management/evaluation-templates/?type={template.evaluation_type}&template={template.id}")

    existing_ids = list(template.criteria.order_by("order", "id").values_list("id", flat=True))
    if sorted(ordered_ids) != sorted(existing_ids):
        messages.error(request, "문항 목록이 변경되어 순서를 저장하지 못했습니다. 새로고침 후 다시 시도해주세요.")
        return redirect(f"/management/evaluation-templates/?type={template.evaluation_type}&template={template.id}")

    for index, criterion_id in enumerate(ordered_ids, start=1):
        EvaluationCriterion.objects.filter(pk=criterion_id, template=template).update(order=index)

    messages.success(request, "평가 문항 순서를 저장했습니다.")
    return redirect(f"/management/evaluation-templates/?type={template.evaluation_type}&template={template.id}")

@admin_required
def admin_missing_evaluations(request):
    """회차별로 아직 제출되지 않은 팀/개인 평가를 평가자-대상 단위로 보여준다."""
    rounds = EvaluationRound.objects.all().order_by("-start_at")
    selected_round = _selected_round(request, rounds)
    evaluation_type = request.GET.get("type", "all")
    query = request.GET.get("q", "").strip().lower()

    rows = []
    summary = {
        "team_missing": 0,
        "personal_missing": 0,
        "draft_count": 0,
        "not_started_count": 0,
        "exempt_team_count": 0,
    }

    if selected_round:
        memberships_qs = list(
            TeamMembership.objects.filter(
                team__evaluation_round=selected_round,
                team__is_active=True,
                student__is_active=True,
            ).select_related("team", "student__user").order_by("student__user__first_name", "student__user__username")
        )

        active_teams = list(
            Team.objects.filter(evaluation_round=selected_round, is_active=True).order_by("name")
        )

        memberships_by_team = {}
        for membership in memberships_qs:
            memberships_by_team.setdefault(membership.team_id, []).append(membership)

        attendance_map = dict(
            RoundAttendance.objects.filter(
                evaluation_round=selected_round,
                student_id__in=[m.student_id for m in memberships_qs],
            ).values_list("student_id", "status")
        )

        team_eval_map = {
            (e.evaluator_id, e.target_team_id): e
            for e in TeamEvaluation.objects.filter(
                evaluation_round=selected_round,
                evaluator__is_active=True,
            )
        }
        personal_eval_map = {
            (e.evaluator_id, e.target_student_id): e
            for e in PersonalEvaluation.objects.filter(
                evaluation_round=selected_round,
                evaluator__is_active=True,
            )
        }

        for membership in memberships_qs:
            evaluator = membership.student
            evaluator_name = evaluator.name
            evaluator_email = evaluator.user.email
            team = membership.team
            attendance_status = attendance_map.get(evaluator.id, RoundAttendance.Status.PRESENT)

            # 발표 당일 출석자만 다른 팀 평가 의무가 있다.
            if attendance_status == RoundAttendance.Status.PRESENT:
                for target_team in active_teams:
                    if target_team.id == team.id:
                        continue
                    evaluation = team_eval_map.get((evaluator.id, target_team.id))
                    if evaluation and evaluation.is_submitted:
                        continue
                    state = "draft" if evaluation else "not_started"
                    summary["team_missing"] += 1
                    summary["draft_count" if state == "draft" else "not_started_count"] += 1
                    rows.append({
                        "type": "team",
                        "type_label": "팀 평가",
                        "evaluator_name": evaluator_name,
                        "evaluator_email": evaluator_email,
                        "evaluator_team": team.name,
                        "target_name": target_team.name,
                        "state": state,
                        "state_label": "임시저장" if state == "draft" else "미시작",
                        "attendance_label": "출석",
                    })
            else:
                summary["exempt_team_count"] += max(len(active_teams) - 1, 0)

            # 개인 평가는 결석/공결 여부와 관계없이 같은 팀원 대상으로 진행한다.
            for target_membership in memberships_by_team.get(team.id, []):
                target = target_membership.student
                if target.id == evaluator.id:
                    continue
                evaluation = personal_eval_map.get((evaluator.id, target.id))
                if evaluation and evaluation.is_submitted:
                    continue
                state = "draft" if evaluation else "not_started"
                summary["personal_missing"] += 1
                summary["draft_count" if state == "draft" else "not_started_count"] += 1
                rows.append({
                    "type": "personal",
                    "type_label": "개인 평가",
                    "evaluator_name": evaluator_name,
                    "evaluator_email": evaluator_email,
                    "evaluator_team": team.name,
                    "target_name": target.name,
                    "state": state,
                    "state_label": "임시저장" if state == "draft" else "미시작",
                    "attendance_label": dict(RoundAttendance.Status.choices).get(attendance_status, "출석"),
                })

        if evaluation_type in {"team", "personal"}:
            rows = [row for row in rows if row["type"] == evaluation_type]

        if query:
            rows = [
                row for row in rows
                if query in row["evaluator_name"].lower()
                or query in (row["evaluator_email"] or "").lower()
                or query in row["evaluator_team"].lower()
                or query in row["target_name"].lower()
            ]

    summary["total_missing"] = summary["team_missing"] + summary["personal_missing"]

    return render(
        request,
        "admin_ui/missing_evaluations.html",
        _base_context(
            rounds=rounds,
            selected_round=selected_round,
            rows=rows,
            summary=summary,
            evaluation_type=evaluation_type,
            query=request.GET.get("q", "").strip(),
        ),
    )

@admin_required
def admin_evaluation_results(request):
    rounds = EvaluationRound.objects.all().order_by("-start_at")
    selected_round = _selected_round(request, rounds)

    if request.method == "POST" and selected_round:
        _recalculate_round_results(selected_round)
        messages.success(request, f"평가 결과를 다시 계산했습니다. 개인 {selected_round.personal_weight}% + 팀 {selected_round.team_weight}% 및 관리자 보정점수가 반영됩니다.")
        return redirect(f"{request.path}?round={selected_round.id}")

    query = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "all")
    result_rows = []
    team_rows = []
    stats = {
        "team_completed": 0,
        "team_required": 0,
        "personal_completed": 0,
        "personal_required": 0,
        "calculated": 0,
        "excluded": 0,
        "publish_status": "미공개",
        "publish_detail": "학생에게 아직 결과를 공개하지 않았습니다.",
        "completion_percent": 0,
        "top_score": None,
        "average_score": None,
    }
    setting = None
    review_flags = []
    review_flag_count = 0

    if selected_round:
        memberships_qs = TeamMembership.objects.filter(
            team__evaluation_round=selected_round,
            team__is_active=True,
            student__is_active=True,
        ).select_related("team", "student__user")
        memberships = {m.student_id: m.team for m in memberships_qs}
        member_count_by_team = {}
        for membership in memberships_qs:
            member_count_by_team[membership.team_id] = member_count_by_team.get(membership.team_id, 0) + 1

        active_teams = list(Team.objects.filter(evaluation_round=selected_round, is_active=True).order_by("name"))
        attendance_map = dict(
            RoundAttendance.objects.filter(
                evaluation_round=selected_round,
                student_id__in=[m.student_id for m in memberships_qs],
            ).values_list("student_id", "status")
        )
        stats.update(_submission_progress(selected_round, memberships_qs, active_teams, attendance_map))
        required_total = stats["team_required"] + stats["personal_required"]
        completed_total = min(stats["team_completed"], stats["team_required"]) + min(
            stats["personal_completed"], stats["personal_required"]
        )
        stats["completion_percent"] = round((completed_total / required_total) * 100) if required_total else 0

        all_review_flags = _evaluation_review_flags(selected_round)
        review_flag_count = len(all_review_flags)
        review_flags = all_review_flags[:12]

        results = StudentResult.objects.filter(evaluation_round=selected_round).select_related("student__user")
        if query:
            results = results.filter(
                Q(student__user__first_name__icontains=query)
                | Q(student__user__last_name__icontains=query)
                | Q(student__user__email__icontains=query)
                | Q(student__user__username__icontains=query)
                | Q(student__team_memberships__team__name__icontains=query)
            ).distinct()
        if status_filter == "included":
            results = results.filter(is_excluded=False)
        elif status_filter == "excluded":
            results = results.filter(is_excluded=True)
        results = results.order_by("is_excluded", "rank", "student__user__first_name", "student__user__username")

        for r in results:
            team = memberships.get(r.student_id)
            result_rows.append({
                "student_name": r.student.name,
                "email": r.student.user.email,
                "team_name": team.name if team else "-",
                "team_score": r.team_score,
                "personal_score": r.personal_score,
                "base_score": r.base_score,
                "adjustment_score": r.adjustment_score,
                "adjustment_reason": r.adjustment_reason,
                "result_id": r.id,
                "final_score": r.final_score,
                "rank": r.rank if not r.is_excluded else None,
                "is_excluded": r.is_excluded,
            })

        all_results = StudentResult.objects.filter(evaluation_round=selected_round)
        included_results = all_results.filter(is_excluded=False)
        stats["calculated"] = included_results.count()
        stats["excluded"] = all_results.filter(is_excluded=True).count()
        score_summary = included_results.aggregate(avg=Avg("final_score"))
        stats["average_score"] = score_summary["avg"]
        top_result = included_results.order_by("rank", "-final_score").select_related("student__user").first()
        if top_result:
            stats["top_score"] = top_result.final_score
            stats["top_student"] = top_result.student.name

        team_eval_counts = dict(
            TeamEvaluation.objects.filter(evaluation_round=selected_round, is_submitted=True)
            .values("target_team_id").annotate(c=Count("id")).values_list("target_team_id", "c")
        )
        team_result_map = {r.team_id: r for r in TeamResult.objects.filter(evaluation_round=selected_round)}
        for team in active_teams:
            team_result = team_result_map.get(team.id)
            team_rows.append({
                "name": team.name,
                "member_count": member_count_by_team.get(team.id, 0),
                "evaluation_count": team_eval_counts.get(team.id, 0),
                "score": team_result.score if team_result else None,
                "rank": team_result.rank if team_result and not team_result.is_excluded else None,
                "is_excluded": team_result.is_excluded if team_result else False,
            })

        setting = ResultPublishSetting.objects.filter(evaluation_round=selected_round).first()
        if setting:
            now = timezone.now()
            if setting.is_published or (setting.publish_at and setting.publish_at <= now):
                stats["publish_status"] = "공개 중"
                stats["publish_detail"] = "학생 결과 화면에서 현재 결과를 확인할 수 있습니다."
            elif setting.publish_at and setting.publish_at > now:
                stats["publish_status"] = "예약 공개"
                stats["publish_detail"] = f"{timezone.localtime(setting.publish_at).strftime('%Y-%m-%d %H:%M')} 공개 예정"

    return render(request, "admin_ui/evaluation_results.html", _base_context(
        rounds=rounds,
        selected_round=selected_round,
        result_rows=result_rows,
        team_rows=team_rows,
        stats=stats,
        setting=setting,
        query=query,
        status_filter=status_filter,
        review_flags=review_flags,
        review_flag_count=review_flag_count,
    ))

@admin_required
@require_POST
def admin_result_weights_save(request):
    round_obj = get_object_or_404(EvaluationRound, id=request.POST.get("round_id"))
    try:
        personal_weight = int(request.POST.get("personal_weight", 60))
        team_weight = int(request.POST.get("team_weight", 40))
    except (TypeError, ValueError):
        messages.error(request, "가중치는 숫자로 입력해 주세요.")
        return _redirect_back(request, "admin_evaluation_results")

    if not (0 <= personal_weight <= 100 and 0 <= team_weight <= 100):
        messages.error(request, "가중치는 0~100 사이여야 합니다.")
        return _redirect_back(request, "admin_evaluation_results")
    if personal_weight + team_weight != 100:
        messages.error(request, "개인 평가와 팀 평가 가중치 합계는 100%여야 합니다.")
        return _redirect_back(request, "admin_evaluation_results")

    round_obj.personal_weight = personal_weight
    round_obj.team_weight = team_weight
    round_obj.save(update_fields=["personal_weight", "team_weight", "updated_at"])
    _recalculate_round_results(round_obj)
    messages.success(request, f"가중치를 개인 {personal_weight}% / 팀 {team_weight}%로 저장하고 결과를 재계산했습니다.")
    return _redirect_back(request, "admin_evaluation_results")

@admin_required
@require_POST
def admin_student_result_adjust(request, result_id):
    result = get_object_or_404(StudentResult.objects.select_related("evaluation_round", "student__user"), id=result_id)
    raw_score = (request.POST.get("adjustment_score") or "0").strip()
    reason = (request.POST.get("adjustment_reason") or "").strip()
    try:
        adjustment = Decimal(raw_score)
    except InvalidOperation:
        messages.error(request, "보정점수는 숫자로 입력해 주세요.")
        return _redirect_back(request, "admin_evaluation_results")

    if adjustment != 0 and not reason:
        messages.error(request, "보정점수를 적용할 때는 사유를 입력해 주세요.")
        return _redirect_back(request, "admin_evaluation_results")

    result.adjustment_score = adjustment
    result.adjustment_reason = reason if adjustment != 0 else ""
    result.save(update_fields=["adjustment_score", "adjustment_reason", "updated_at"])
    _recalculate_round_results(result.evaluation_round)
    messages.success(request, f"{result.student.name} 학생의 관리자 보정점수를 저장했습니다.")
    return _redirect_back(request, "admin_evaluation_results")

@admin_required
def admin_evaluation_results_excel_export(request):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from urllib.parse import quote

    rounds = EvaluationRound.objects.all().order_by("-start_at")
    selected_round = _selected_round(request, rounds)
    if not selected_round:
        messages.error(request, "내보낼 평가 회차가 없습니다.")
        return redirect("admin_evaluation_results")

    # 내보내기 직전에 최신 제출 내용을 다시 집계한다.
    _recalculate_round_results(selected_round)
    memberships = {
        m.student_id: m.team
        for m in TeamMembership.objects.filter(team__evaluation_round=selected_round).select_related("team")
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "종합 결과"
    headers = ["순위", "수강생", "이메일", "팀", "팀 점수", "개인 점수", "가중 합산", "관리자 보정", "보정 사유", "최종 점수", "집계 상태"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="E8EEF8")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    results = StudentResult.objects.filter(evaluation_round=selected_round).select_related("student__user").order_by(
        "is_excluded", "rank", "student__user__first_name", "student__user__username"
    )
    for result in results:
        team = memberships.get(result.student_id)
        ws.append([
            result.rank if not result.is_excluded else "제외",
            result.student.name,
            result.student.user.email,
            team.name if team else "-",
            float(result.team_score),
            float(result.personal_score),
            float(result.base_score),
            float(result.adjustment_score),
            result.adjustment_reason,
            float(result.final_score),
            "집계 제외" if result.is_excluded else "정상",
        ])

    team_ws = wb.create_sheet("팀 결과")
    team_ws.append(["순위", "팀", "인원", "제출 평가 수", "평균 점수", "집계 상태"])
    for cell in team_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    member_counts = dict(
        TeamMembership.objects.filter(team__evaluation_round=selected_round)
        .values("team_id").annotate(c=Count("id")).values_list("team_id", "c")
    )
    eval_counts = dict(
        TeamEvaluation.objects.filter(evaluation_round=selected_round, is_submitted=True)
        .values("target_team_id").annotate(c=Count("id")).values_list("target_team_id", "c")
    )
    result_map = {r.team_id: r for r in TeamResult.objects.filter(evaluation_round=selected_round)}
    for team in Team.objects.filter(evaluation_round=selected_round, is_active=True).order_by("name"):
        result = result_map.get(team.id)
        team_ws.append([
            result.rank if result and not result.is_excluded else ("제외" if result and result.is_excluded else "-"),
            team.name,
            member_counts.get(team.id, 0),
            eval_counts.get(team.id, 0),
            float(result.score) if result else None,
            "집계 제외" if result and result.is_excluded else "정상",
        ])

    for sheet in (ws, team_ws):
        for column in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 3, 12), 32)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{selected_round.name}_평가결과.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response

@admin_required
def admin_team_scores(request):
    rounds = EvaluationRound.objects.all().order_by("-start_at")
    selected_round = _selected_round(request, rounds)
    if request.method == "POST" and selected_round:
        _recalculate_round_results(selected_round)
        messages.success(request, "팀 점수를 다시 계산했습니다.")
        return redirect(f"{request.path}?round={selected_round.id}")

    rows = []
    if selected_round:
        counts = dict(
            TeamEvaluation.objects.filter(evaluation_round=selected_round, is_submitted=True)
            .values("target_team_id").annotate(c=Count("id")).values_list("target_team_id", "c")
        )
        results = {r.team_id: r for r in TeamResult.objects.filter(evaluation_round=selected_round)}
        for team in Team.objects.filter(evaluation_round=selected_round, is_active=True).order_by("name"):
            result = results.get(team.id)
            rows.append({
                "name": team.name, "evaluation_count": counts.get(team.id, 0),
                "score": result.score if result else None, "rank": result.rank if result and not result.is_excluded else None,
                "is_excluded": result.is_excluded if result else False,
            })
    return render(request, "admin_ui/team_scores.html", _base_context(rounds=rounds, selected_round=selected_round, team_scores=rows))

@admin_required
def admin_personal_scores(request):
    rounds = EvaluationRound.objects.all().order_by("-start_at")
    selected_round = _selected_round(request, rounds)
    if request.method == "POST" and selected_round:
        _recalculate_round_results(selected_round)
        messages.success(request, "개인 점수를 다시 계산했습니다.")
        return redirect(f"{request.path}?round={selected_round.id}")

    rows = []
    if selected_round:
        counts = dict(
            PersonalEvaluation.objects.filter(evaluation_round=selected_round, is_submitted=True)
            .values("target_student_id").annotate(c=Count("id")).values_list("target_student_id", "c")
        )
        for r in StudentResult.objects.filter(evaluation_round=selected_round).select_related("student__user").order_by("student__user__first_name"):
            rows.append({
                "student_name": r.student.name, "evaluation_count": counts.get(r.student_id, 0),
                "score": r.personal_score, "team_score": r.team_score, "final_score": r.final_score,
                "is_excluded": r.is_excluded,
            })
    return render(request, "admin_ui/personal_scores.html", _base_context(rounds=rounds, selected_round=selected_round, personal_scores=rows))

@admin_required
def admin_rankings(request):
    rounds = EvaluationRound.objects.all().order_by("-start_at")
    selected_round = _selected_round(request, rounds)
    if request.method == "POST" and selected_round:
        _recalculate_round_results(selected_round)
        messages.success(request, "순위를 다시 계산했습니다.")
        return redirect(f"{request.path}?round={selected_round.id}")

    rankings = []
    if selected_round:
        memberships = {m.student_id: m.team for m in TeamMembership.objects.filter(team__evaluation_round=selected_round).select_related("team")}
        results = StudentResult.objects.filter(evaluation_round=selected_round).select_related("student__user").order_by("is_excluded", "rank", "student__user__first_name")
        for r in results:
            team = memberships.get(r.student_id)
            rankings.append({
                "rank": r.rank if not r.is_excluded else "제외", "student_name": r.student.name,
                "team_name": team.name if team else "-", "team_score": r.team_score,
                "personal_score": r.personal_score, "final_score": r.final_score, "is_excluded": r.is_excluded,
            })
    return render(request, "admin_ui/rankings.html", _base_context(rounds=rounds, selected_round=selected_round, rankings=rankings))

@admin_required
def admin_result_settings(request):
    rounds = EvaluationRound.objects.all().order_by("-start_at")
    selected_round = _selected_round(request, rounds)
    setting = None
    if selected_round:
        setting, _ = ResultPublishSetting.objects.get_or_create(evaluation_round=selected_round)

    if request.method == "POST" and selected_round and setting:
        action = request.POST.get("action", "save")
        setting.show_team_first_place = "show_team_first_place" in request.POST
        setting.show_all_team_ranks = "show_all_team_ranks" in request.POST
        setting.show_personal_score = "show_personal_score" in request.POST
        setting.show_overall_rank = "show_overall_rank" in request.POST
        setting.show_comments = "show_comments" in request.POST

        publish_at_raw = request.POST.get("publish_at", "").strip()
        if publish_at_raw:
            from datetime import datetime
            try:
                dt = datetime.fromisoformat(publish_at_raw)
                setting.publish_at = timezone.make_aware(dt) if timezone.is_naive(dt) else dt
            except ValueError:
                setting.publish_at = None
        else:
            setting.publish_at = None

        if action == "publish":
            _recalculate_round_results(selected_round)
            setting.is_published = True
            setting.publish_at = timezone.now()
            messages.success(request, "결과를 즉시 공개했습니다.")
        elif action == "unpublish":
            setting.is_published = False
            setting.publish_at = None
            messages.success(request, "결과 공개를 중지했습니다.")
        else:
            setting.is_published = False if setting.publish_at and setting.publish_at > timezone.now() else setting.is_published
            messages.success(request, "결과 공개 설정을 저장했습니다.")
        setting.save()
        return redirect(f"{request.path}?round={selected_round.id}")

    publish_options = []
    if setting:
        publish_options = [
            {"key": "show_team_first_place", "label": "팀 1위 공개", "enabled": setting.show_team_first_place},
            {"key": "show_all_team_ranks", "label": "전체 팀 순위 공개", "enabled": setting.show_all_team_ranks},
            {"key": "show_personal_score", "label": "개인 점수 공개", "enabled": setting.show_personal_score},
            {"key": "show_overall_rank", "label": "개인 종합 순위 공개", "enabled": setting.show_overall_rank},
            {"key": "show_comments", "label": "평가 코멘트 공개", "enabled": setting.show_comments},
        ]
    effective_published = bool(setting and (setting.is_published or (setting.publish_at and setting.publish_at <= timezone.now())))
    return render(request, "admin_ui/result_settings.html", _base_context(
        rounds=rounds, selected_round=selected_round, setting=setting, publish_options=publish_options, effective_published=effective_published
    ))

@admin_required
def admin_seed_management(request):
    rounds = EvaluationRound.objects.all().order_by("-start_at")
    selected_round = _selected_round(request, rounds)

    if request.method == "POST":
        weight_round = get_object_or_404(EvaluationRound, id=request.POST.get("weight_round_id"))
        try:
            seed_weight = int(request.POST.get("seed_weight", 100))
            seed_team_weight = int(request.POST.get("seed_team_weight", 40))
            seed_personal_weight = int(request.POST.get("seed_personal_weight", 60))
        except (TypeError, ValueError):
            messages.error(request, "Seed 설정값은 숫자로 입력해 주세요.")
            return _redirect_back(request, "admin_seed_management")

        if not 0 <= seed_weight <= 100:
            messages.error(request, "회차 반영 가중치는 0~100 사이여야 합니다.")
            return _redirect_back(request, "admin_seed_management")
        if not 0 <= seed_team_weight <= 100 or not 0 <= seed_personal_weight <= 100:
            messages.error(request, "팀/개인 Seed 비율은 각각 0~100 사이여야 합니다.")
            return _redirect_back(request, "admin_seed_management")
        if seed_team_weight + seed_personal_weight != 100:
            messages.error(request, "팀 Seed 비율과 개인 Seed 비율의 합은 100이어야 합니다.")
            return _redirect_back(request, "admin_seed_management")

        weight_round.seed_weight = seed_weight
        weight_round.seed_team_weight = seed_team_weight
        weight_round.seed_personal_weight = seed_personal_weight
        weight_round.save(
            update_fields=[
                "seed_weight",
                "seed_team_weight",
                "seed_personal_weight",
                "updated_at",
            ]
        )
        messages.success(
            request,
            f"{weight_round.name}: 회차 {seed_weight}% · 팀 {seed_team_weight}% · 개인 {seed_personal_weight}%로 저장했습니다.",
        )
        return _redirect_back(request, "admin_seed_management")

    seed_rows = []
    history_rounds = []
    selected_round_results = []

    if selected_round:
        history_rounds = list(
            EvaluationRound.objects.filter(
                start_at__lt=selected_round.start_at,
                status=EvaluationRound.Status.ENDED,
            ).order_by("-start_at", "-id")
        )

        seed_scores = _cumulative_seed_scores_before(selected_round)
        students = (
            Student.objects.filter(id__in=seed_scores.keys())
            .select_related("user")
            .order_by("user__first_name", "user__username")
        )
        for student in students:
            seed_rows.append({
                "student_name": student.name,
                "email": student.user.email,
                "seed_score": seed_scores.get(student.id),
            })
        seed_rows.sort(key=lambda row: float(row["seed_score"] or 0), reverse=True)
        for index, row in enumerate(seed_rows, start=1):
            row["seed_rank"] = index

        selected_round_results = list(
            StudentResult.objects.filter(
                evaluation_round=selected_round,
                is_excluded=False,
            ).select_related("student__user").order_by("rank", "-final_score")
        )

    return render(
        request,
        "admin_ui/seed_management.html",
        _base_context(
            rounds=rounds,
            selected_round=selected_round,
            history_rounds=history_rounds,
            seed_rows=seed_rows,
            selected_round_results=selected_round_results,
        ),
    )
