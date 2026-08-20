"""Student Excel import/export views."""

from io import BytesIO

from django.contrib import messages
from django.contrib.auth.models import User
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.http import HttpResponse

from .admin_skills import _sync_student_common_skills
from .common import (
    _current_round,
    _excel_bool,
    _normalize_excel_header,
    _redirect_back,
    _round_teams,
    _sync_student_team,
    admin_required,
)
from ..models import Student, TeamMembership


@admin_required
def admin_students_excel_upload(request):
    """Excel(.xlsx)로 수강생을 일괄 등록/수정한다. 이메일을 기준으로 기존 계정을 갱신한다."""
    uploaded = request.FILES.get("excel_file")
    if not uploaded:
        messages.error(request, "업로드할 Excel 파일을 선택해주세요.")
        return _redirect_back(request, "admin_students")

    if not uploaded.name.lower().endswith(".xlsx"):
        messages.error(request, ".xlsx 형식의 Excel 파일만 업로드할 수 있습니다.")
        return _redirect_back(request, "admin_students")

    if uploaded.size > 5 * 1024 * 1024:
        messages.error(request, "Excel 파일은 5MB 이하만 업로드할 수 있습니다.")
        return _redirect_back(request, "admin_students")

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(uploaded, read_only=True, data_only=True)
        worksheet = workbook.active
    except Exception:
        messages.error(request, "Excel 파일을 읽을 수 없습니다. 파일이 손상되지 않았는지 확인해주세요.")
        return _redirect_back(request, "admin_students")

    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        messages.error(request, "Excel 파일에 데이터가 없습니다.")
        return _redirect_back(request, "admin_students")

    aliases = {
        "name": {"이름", "성명", "name", "studentname"},
        "email": {"이메일", "메일", "email", "emailaddress"},
        "password": {"비밀번호", "임시비밀번호", "password", "pw"},
        "affiliation": {"소속", "전공", "affiliation", "major", "department", "학과"},
        "team": {"팀", "팀명", "소속팀", "team", "teamname"},
        "is_active": {"활성", "상태", "활성상태", "isactive", "active"},
    }
    normalized_aliases = {
        key: {_normalize_excel_header(alias) for alias in values}
        for key, values in aliases.items()
    }
    headers = [_normalize_excel_header(value) for value in raw_headers]
    columns = {}
    for key, candidates in normalized_aliases.items():
        for index, header in enumerate(headers):
            if header in candidates:
                columns[key] = index
                break

    if "name" not in columns or "email" not in columns:
        messages.error(request, "Excel 첫 행에 '이름'과 '이메일' 열이 반드시 있어야 합니다.")
        return _redirect_back(request, "admin_students")

    current_round = _current_round()
    team_map = {
        team.name.strip().lower(): team
        for team in _round_teams(current_round)
    } if current_round else {}

    created = 0
    updated = 0
    skipped = 0
    errors = []

    def cell(row, key):
        index = columns.get(key)
        if index is None or index >= len(row):
            return ""
        value = row[index]
        return "" if value is None else str(value).strip()

    for excel_row, row in enumerate(rows, start=2):
        if not any(value not in (None, "") for value in row):
            continue

        name = cell(row, "name")
        email = cell(row, "email").lower()
        password = cell(row, "password")
        affiliation = cell(row, "affiliation")
        team_name = cell(row, "team")
        active_raw = cell(row, "is_active")

        try:
            if not name or not email:
                raise ValueError("이름과 이메일은 필수입니다.")
            if "@" not in email:
                raise ValueError("이메일 형식이 올바르지 않습니다.")

            active = _excel_bool(active_raw, default=True) if "is_active" in columns else None
            team = None
            if "team" in columns and team_name:
                if not current_round:
                    raise ValueError("현재 평가 회차가 없어 팀을 배정할 수 없습니다.")
                team = team_map.get(team_name.lower())
                if not team:
                    raise ValueError(f"현재 회차에서 '{team_name}' 팀을 찾을 수 없습니다.")

            with transaction.atomic():
                user = User.objects.filter(Q(email__iexact=email) | Q(username__iexact=email)).first()
                if user:
                    student = getattr(user, "student_profile", None)
                    if student is None:
                        raise ValueError("같은 이메일의 관리자/일반 계정이 이미 존재합니다.")
                    user.first_name = name
                    user.last_name = ""
                    user.email = email
                    user.username = email
                    if active is not None:
                        user.is_active = active
                    if password:
                        if len(password) < 8:
                            raise ValueError("비밀번호는 8자 이상이어야 합니다.")
                        user.set_password(password)
                    user.save()
                    if "affiliation" in columns:
                        student.affiliation = affiliation
                    if active is not None:
                        student.is_active = active
                    student.save()
                    if "team" in columns:
                        _sync_student_team(student, team, current_round)
                    updated += 1
                else:
                    if not password:
                        raise ValueError("신규 수강생은 임시 비밀번호가 필요합니다.")
                    if len(password) < 8:
                        raise ValueError("비밀번호는 8자 이상이어야 합니다.")
                    new_active = True if active is None else active
                    user = User.objects.create_user(
                        username=email,
                        email=email,
                        password=password,
                        first_name=name,
                        is_active=new_active,
                    )
                    student = Student.objects.create(user=user, affiliation=affiliation, is_active=new_active)
                    _sync_student_common_skills(student)
                    if "team" in columns:
                        _sync_student_team(student, team, current_round)
                    created += 1
        except (ValueError, IntegrityError) as exc:
            skipped += 1
            errors.append(f"{excel_row}행: {exc}")
        except Exception as exc:
            skipped += 1
            errors.append(f"{excel_row}행: 처리 중 오류가 발생했습니다. ({exc})")

    workbook.close()
    messages.success(request, f"Excel 업로드 완료: 신규 {created}명, 수정 {updated}명, 제외 {skipped}명")
    if errors:
        preview = " / ".join(errors[:5])
        if len(errors) > 5:
            preview += f" / 외 {len(errors) - 5}건"
        messages.warning(request, preview)
    return _redirect_back(request, "admin_students")


@admin_required
def admin_students_excel_export(request):
    """현재 수강생 목록을 다시 업로드 가능한 .xlsx 형식으로 내려준다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    current_round = _current_round()
    students = list(Student.objects.select_related("user").order_by("user__first_name", "user__username"))
    membership_map = {}
    if current_round and students:
        memberships = TeamMembership.objects.filter(
            student_id__in=[student.id for student in students],
            team__evaluation_round=current_round,
        ).select_related("team")
        membership_map = {membership.student_id: membership.team.name for membership in memberships}

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "수강생"
    headers = ["이름", "이메일", "임시비밀번호", "소속", "팀명", "활성상태"]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    for student in students:
        worksheet.append([
            student.name,
            student.email,
            "",
            student.affiliation,
            membership_map.get(student.id, ""),
            "활성" if student.is_active and student.user.is_active else "비활성",
        ])

    widths = {"A": 18, "B": 30, "C": 18, "D": 20, "E": 18, "F": 12}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="students.xlsx"'
    return response
