"""Admin team list, export and CRUD views."""

from io import BytesIO

from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from .common import _base_context, _current_round, _redirect_back, admin_required
from ..forms import TeamForm
from ..models import EvaluationRound, Student, Team, TeamMembership


@admin_required
def admin_teams(request):
    selected_round_id = request.GET.get("round", "").strip()
    query = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "all").strip() or "all"
    rounds = EvaluationRound.objects.all().order_by("-start_at")
    selected_round = rounds.filter(pk=selected_round_id).first() if selected_round_id else None
    if selected_round is None:
        selected_round = _current_round()

    base_teams_qs = (
        Team.objects.filter(evaluation_round=selected_round)
        if selected_round
        else Team.objects.none()
    )
    all_round_teams = list(base_teams_qs.order_by("name"))

    teams_qs = base_teams_qs
    if query:
        teams_qs = teams_qs.filter(
            Q(name__icontains=query)
            | Q(project_title__icontains=query)
            | Q(memberships__student__user__first_name__icontains=query)
            | Q(memberships__student__user__email__icontains=query)
        ).distinct()
    if status_filter == "active":
        teams_qs = teams_qs.filter(is_active=True)
    elif status_filter == "inactive":
        teams_qs = teams_qs.filter(is_active=False)
    teams = list(teams_qs.order_by("name"))

    display_team_ids = [team.id for team in teams]
    memberships = (
        TeamMembership.objects.filter(team_id__in=display_team_ids)
        .select_related("student__user", "team")
        .order_by("team__name", "student__user__first_name", "student__user__username")
    )
    member_map = {team.id: [] for team in teams}
    for membership in memberships:
        member_map.setdefault(membership.team_id, []).append(membership.student)
    for team in teams:
        team.members = member_map.get(team.id, [])
        team.member_count = len(team.members)
        team.status_display = "활성" if team.is_active else "비활성"
        team.leader = team.members[0] if team.members else None

    active_students = Student.objects.filter(is_active=True, user__is_active=True)
    assigned_ids = set()
    if selected_round:
        round_memberships = TeamMembership.objects.filter(team__evaluation_round=selected_round)
        assigned_ids = set(round_memberships.values_list("student_id", flat=True))

    total_team_count = len(all_round_teams)
    assigned_count = len(assigned_ids)
    average_members = round(assigned_count / total_team_count, 1) if total_team_count else 0
    unassigned_students = list(
        active_students.exclude(id__in=assigned_ids)
        .select_related("user")
        .order_by("user__first_name", "user__username")[:8]
    )
    stats = {
        "team_count": total_team_count,
        "assigned_count": assigned_count,
        "active_student_count": active_students.count(),
        "average_members": average_members,
        "unassigned_count": active_students.exclude(id__in=assigned_ids).count(),
        "active_team_count": sum(1 for team in all_round_teams if team.is_active),
    }
    return render(
        request,
        "admin_ui/teams.html",
        _base_context(
            teams=teams,
            rounds=rounds,
            selected_round=selected_round,
            query=query,
            status_filter=status_filter,
            unassigned_students=unassigned_students,
            stats=stats,
        ),
    )


@admin_required
def admin_teams_excel_export(request):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    round_id = request.GET.get("round", "").strip()
    evaluation_round = (
        EvaluationRound.objects.filter(pk=round_id).first()
        if round_id
        else _current_round()
    )
    if not evaluation_round:
        messages.error(request, "내보낼 평가 회차가 없습니다.")
        return redirect("admin_teams")

    teams = list(Team.objects.filter(evaluation_round=evaluation_round).order_by("name"))
    memberships = (
        TeamMembership.objects.filter(team__in=teams)
        .select_related("team", "student__user")
        .order_by("team__name", "student__user__first_name", "student__user__username")
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "팀 구성"
    worksheet.append(["회차", "팀명", "프로젝트", "상태", "학생명", "이메일", "소속"])
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    member_map = {team.id: [] for team in teams}
    for membership in memberships:
        member_map.setdefault(membership.team_id, []).append(membership.student)

    for team in teams:
        members = member_map.get(team.id, [])
        if not members:
            worksheet.append([
                evaluation_round.name,
                team.name,
                team.project_title or "",
                "활성" if team.is_active else "비활성",
                "",
                "",
                "",
            ])
            continue
        for student in members:
            worksheet.append([
                evaluation_round.name,
                team.name,
                team.project_title or "",
                "활성" if team.is_active else "비활성",
                student.user.get_full_name() or student.user.username,
                student.user.email,
                student.affiliation or "",
            ])

    for column, width in {"A": 22, "B": 14, "C": 30, "D": 12, "E": 16, "F": 30, "G": 18}.items():
        worksheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="teams_round_{evaluation_round.id}.xlsx"'
    return response


@admin_required
@require_POST
@transaction.atomic
def admin_team_create(request):
    form = TeamForm(request.POST, rounds=EvaluationRound.objects.all())
    if not form.is_valid():
        for errors in form.errors.values():
            for error in errors:
                messages.error(request, error)
        return _redirect_back(request, "admin_teams")
    team = form.save()
    messages.success(request, f"{team.name}을(를) 생성했습니다.")
    return _redirect_back(request, f"/management/teams/?round={team.evaluation_round_id}")


@admin_required
@require_POST
@transaction.atomic
def admin_team_update(request, team_id):
    team = get_object_or_404(Team, pk=team_id)
    form = TeamForm(request.POST, instance=team, rounds=EvaluationRound.objects.all())
    if not form.is_valid():
        for errors in form.errors.values():
            for error in errors:
                messages.error(request, error)
        return _redirect_back(request, f"/management/teams/?round={team.evaluation_round_id}")
    team = form.save()
    messages.success(request, f"{team.name} 정보를 수정했습니다.")
    return _redirect_back(request, f"/management/teams/?round={team.evaluation_round_id}")


@admin_required
@require_POST
@transaction.atomic
def admin_team_delete(request, team_id):
    team = get_object_or_404(Team, pk=team_id)
    round_id = team.evaluation_round_id
    name = team.name
    if team.memberships.exists():
        messages.error(
            request,
            "팀원이 배정된 팀은 바로 삭제할 수 없습니다. 먼저 팀원을 이동하거나 배정을 해제하세요.",
        )
        return _redirect_back(request, f"/management/teams/?round={round_id}")
    team.delete()
    messages.success(request, f"{name}을(를) 삭제했습니다.")
    return _redirect_back(request, f"/management/teams/?round={round_id}")
