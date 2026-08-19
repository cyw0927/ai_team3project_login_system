import uuid

from .common import *

def _sync_student_common_skills(student):
    """공통 역량 사전에 있는 모든 역량을 학생에게 0점으로 보장한다."""
    existing_skill_ids = set(
        StudentSkill.objects.filter(student=student).values_list("skill_id", flat=True)
    )
    missing = [
        StudentSkill(student=student, skill=skill, score=0)
        for skill in Skill.objects.all()
        if skill.id not in existing_skill_ids
    ]
    if missing:
        StudentSkill.objects.bulk_create(missing, ignore_conflicts=True)
    return len(missing)


def _sync_skill_to_students(skill):
    """새 공통 역량을 모든 활성 수강생에게 0점으로 배포한다."""
    student_ids = list(
        Student.objects.filter(is_active=True, user__is_active=True)
        .values_list("id", flat=True)
    )
    existing_student_ids = set(
        StudentSkill.objects.filter(skill=skill, student_id__in=student_ids)
        .values_list("student_id", flat=True)
    )
    missing = [
        StudentSkill(student_id=student_id, skill=skill, score=0)
        for student_id in student_ids
        if student_id not in existing_student_ids
    ]
    if missing:
        StudentSkill.objects.bulk_create(missing, ignore_conflicts=True)
    return len(missing)


@admin_required
def admin_skills(request):
    """공통 역량 사전과 수강생 적용 현황을 관리한다."""
    skills = list(
        Skill.objects.annotate(
            profile_count=Count("student_profiles", distinct=True),
            assignment_count=Count("assignment_requirements", distinct=True),
            growth_task_count=Count("required_by_tasks", distinct=True),
        ).order_by("name")
    )
    active_student_count = Student.objects.filter(
        is_active=True,
        user__is_active=True,
    ).count()

    for skill in skills:
        skill.coverage_percent = (
            round((skill.profile_count / active_student_count) * 100)
            if active_student_count else 0
        )

    return render(
        request,
        "admin_ui/skills.html",
        _base_context(
            skills=skills,
            active_student_count=active_student_count,
        ),
    )


@admin_required
@require_POST
@transaction.atomic
def admin_skill_create(request):
    name = (request.POST.get("name") or "").strip()
    description = (request.POST.get("description") or "").strip()

    if not name:
        messages.error(request, "역량명을 입력해주세요.")
        return redirect("admin_skills")
    if len(name) > 80:
        messages.error(request, "역량명은 80자 이하로 입력해주세요.")
        return redirect("admin_skills")
    if len(description) > 240:
        messages.error(request, "설명은 240자 이하로 입력해주세요.")
        return redirect("admin_skills")
    if Skill.objects.filter(name__iexact=name).exists():
        messages.error(request, "같은 이름의 역량이 이미 있습니다.")
        return redirect("admin_skills")

    skill = Skill.objects.create(name=name, description=description)
    applied = _sync_skill_to_students(skill)
    messages.success(
        request,
        f"{skill.name} 역량을 만들고 활성 수강생 {applied}명에게 0점으로 적용했습니다.",
    )
    return redirect("admin_skills")


@admin_required
@require_POST
@transaction.atomic
def admin_skill_update(request, skill_id):
    skill = get_object_or_404(Skill, pk=skill_id)
    name = (request.POST.get("name") or "").strip()
    description = (request.POST.get("description") or "").strip()

    if not name:
        messages.error(request, "역량명을 입력해주세요.")
        return redirect("admin_skills")
    if len(name) > 80 or len(description) > 240:
        messages.error(request, "역량명 또는 설명 길이를 확인해주세요.")
        return redirect("admin_skills")
    if Skill.objects.filter(name__iexact=name).exclude(pk=skill.id).exists():
        messages.error(request, "같은 이름의 역량이 이미 있습니다.")
        return redirect("admin_skills")

    skill.name = name
    skill.description = description
    skill.save(update_fields=["name", "description", "updated_at"])
    applied = _sync_skill_to_students(skill)
    messages.success(
        request,
        f"{skill.name} 역량을 수정했습니다. 누락된 수강생 {applied}명도 0점으로 동기화했습니다.",
    )
    return redirect("admin_skills")


@admin_required
@require_POST
@transaction.atomic
def admin_skill_sync_all(request):
    applied = 0
    active_students = Student.objects.filter(
        is_active=True,
        user__is_active=True,
    ).select_related("user")
    for student in active_students:
        applied += _sync_student_common_skills(student)

    messages.success(
        request,
        f"공통 역량 동기화를 완료했습니다. 누락된 프로필 {applied}개를 0점으로 생성했습니다.",
    )
    return redirect("admin_skills")


@admin_required
@require_POST
@transaction.atomic
def admin_skill_delete(request, skill_id):
    """공통 역량을 실제 삭제한다. 과제에서 사용 중인 역량만 안전상 차단한다."""
    skill = get_object_or_404(Skill, pk=skill_id)
    assignment_count = AssignmentSkill.objects.filter(skill_id=skill.id).count()
    growth_task_count = HRTaskSkill.objects.filter(skill_id=skill.id).count()

    if assignment_count or growth_task_count:
        messages.error(
            request,
            f"{skill.name} 역량은 기본 과제 {assignment_count}개 / 추가 성장과제 {growth_task_count}개에서 사용 중이라 삭제할 수 없습니다.",
        )
        return redirect("admin_skills")

    name = skill.name
    profile_count = StudentSkill.objects.filter(skill_id=skill.id).count()

    deleted_count, _ = Skill.objects.filter(pk=skill.id).delete()

    if deleted_count <= 0 or Skill.objects.filter(pk=skill.id).exists():
        messages.error(request, f"{name} 역량 삭제에 실패했습니다. 다시 시도해주세요.")
        return redirect("admin_skills")

    messages.success(
        request,
        f"{name} 역량을 삭제했습니다. 연결된 수강생 프로필 {profile_count}개도 함께 정리했습니다.",
    )
    return redirect("admin_skills")


@admin_required
def admin_students(request):
    query = request.GET.get("q", "").strip()
    team_filter = request.GET.get("team", "").strip()
    current_round = _current_round()
    teams = _round_teams(current_round)

    students = Student.objects.select_related("user").order_by("user__first_name", "user__username")
    if current_round and team_filter:
        if team_filter == "unassigned":
            students = students.exclude(
                team_memberships__team__evaluation_round=current_round
            )
        elif team_filter.isdigit():
            students = students.filter(
                team_memberships__team__evaluation_round=current_round,
                team_memberships__team_id=int(team_filter),
            )

    if query:
        students = students.filter(
            Q(user__first_name__icontains=query)
            | Q(user__last_name__icontains=query)
            | Q(user__email__icontains=query)
            | Q(user__username__icontains=query)
            | Q(affiliation__icontains=query)
        )

    students = list(students)

    membership_map = {}
    if current_round and students:
        memberships = (
            TeamMembership.objects.filter(
                student_id__in=[student.id for student in students],
                team__evaluation_round=current_round,
            )
            .select_related("team")
            .order_by("team__name")
        )
        membership_map = {membership.student_id: membership.team for membership in memberships}

    badge_map = {}
    if students:
        badge_rows = (
            StudentBadge.objects.filter(student_id__in=[student.id for student in students])
            .select_related("evaluation_round")
            .order_by("student_id", "-evaluation_round__start_at", "badge_type")
        )
        grouped = {}
        for badge in badge_rows:
            grouped.setdefault(badge.student_id, []).append(badge)

        for student_id, badges in grouped.items():
            by_type = {}
            for badge in badges:
                item = by_type.setdefault(
                    badge.badge_type,
                    {
                        "type": badge.badge_type,
                        "label": badge.get_badge_type_display(),
                        "count": 0,
                        "round_names": [],
                    },
                )
                item["count"] += 1
                item["round_names"].append(badge.evaluation_round.name)
            badge_map[student_id] = list(by_type.values())

    missing_student_count = 0
    for student in students:
        student.current_team = membership_map.get(student.id)
        student.badge_summary = badge_map.get(student.id, [])
        student.evaluation_missing_count = 0
        student.has_missing_evaluation = False
        if current_round and student.current_team and student.is_active and student.user.is_active:
            progress = _student_progress(student, current_round, student.current_team)
            required_total = progress["team_total"] + progress["personal_total"]
            completed_total = progress["team_completed"] + progress["personal_completed"]
            student.evaluation_missing_count = max(required_total - completed_total, 0)
            student.has_missing_evaluation = student.evaluation_missing_count > 0
            if student.has_missing_evaluation:
                missing_student_count += 1

    all_students = Student.objects.all()
    stats = {
        "total": all_students.count(),
        "active": all_students.filter(is_active=True, user__is_active=True).count(),
        "inactive": all_students.filter(Q(is_active=False) | Q(user__is_active=False)).distinct().count(),
        "unassigned": (
            all_students.exclude(
                team_memberships__team__evaluation_round=current_round
            ).count()
            if current_round
            else all_students.count()
        ),
    }

    return render(
        request,
        "admin_ui/students.html",
        _base_context(
            students=students,
            teams=teams,
            current_round=current_round,
            query=query,
            team_filter=team_filter,
            stats=stats,
            missing_student_count=missing_student_count,
            skills=Skill.objects.all().order_by("name"),
        ),
    )

@admin_required
@require_POST
@transaction.atomic
def admin_students_bulk_skill_save(request):
    """선택 수강생 또는 전체 활성 수강생에게 여러 역량 점수를 한 번에 입력한다."""
    target_mode = (request.POST.get("target_mode") or "selected").strip()
    selected_ids = [
        int(value)
        for value in request.POST.getlist("student_ids")
        if str(value).isdigit()
    ]

    if target_mode == "all_active":
        target_students = list(
            Student.objects.filter(is_active=True, user__is_active=True)
            .select_related("user")
            .order_by("user__first_name", "user__username")
        )
    else:
        target_students = list(
            Student.objects.filter(pk__in=selected_ids)
            .select_related("user")
            .order_by("user__first_name", "user__username")
        )

    if not target_students:
        messages.error(request, "역량을 적용할 수강생을 선택해주세요.")
        return _redirect_back(request, "admin_students")

    skill_ids = request.POST.getlist("skill_id")
    scores = request.POST.getlist("skill_score")
    notes = request.POST.getlist("skill_note")

    rows = []
    used_skill_ids = set()
    errors = []

    for index, raw_skill_id in enumerate(skill_ids):
        raw_skill_id = (raw_skill_id or "").strip()
        raw_score = (scores[index] if index < len(scores) else "").strip()
        note = (notes[index] if index < len(notes) else "").strip()

        if not raw_skill_id and not raw_score and not note:
            continue
        if not raw_skill_id:
            errors.append(f"{index + 1}번째 행의 역량을 선택해주세요.")
            continue
        if not raw_skill_id.isdigit():
            errors.append(f"{index + 1}번째 행의 역량값이 올바르지 않습니다.")
            continue

        skill_id = int(raw_skill_id)
        if skill_id in used_skill_ids:
            errors.append(f"{index + 1}번째 행에 같은 역량이 중복되었습니다.")
            continue
        used_skill_ids.add(skill_id)

        try:
            score = int(raw_score)
        except (TypeError, ValueError):
            errors.append(f"{index + 1}번째 행의 점수를 입력해주세요.")
            continue

        if not 0 <= score <= 100:
            errors.append(f"{index + 1}번째 행의 점수는 0~100 사이여야 합니다.")
            continue
        if len(note) > 300:
            errors.append(f"{index + 1}번째 행의 메모는 300자 이하로 입력해주세요.")
            continue

        skill = Skill.objects.filter(pk=skill_id).first()
        if not skill:
            errors.append(f"{index + 1}번째 행의 역량을 찾을 수 없습니다.")
            continue

        rows.append((skill, score, note))

    if errors:
        for error in errors[:6]:
            messages.error(request, error)
        if len(errors) > 6:
            messages.error(request, f"외 {len(errors) - 6}건의 입력 오류가 있습니다.")
        return _redirect_back(request, "admin_students")

    if not rows:
        messages.error(request, "입력할 역량을 한 개 이상 추가해주세요.")
        return _redirect_back(request, "admin_students")

    created_count = 0
    updated_count = 0
    for student in target_students:
        for skill, score, note in rows:
            _, created = StudentSkill.objects.update_or_create(
                student=student,
                skill=skill,
                defaults={"score": score, "note": note},
            )
            if created:
                created_count += 1
            else:
                updated_count += 1

    target_label = "전체 활성 수강생" if target_mode == "all_active" else "선택 수강생"
    messages.success(
        request,
        f"{target_label} {len(target_students)}명에게 역량 {len(rows)}개를 일괄 적용했습니다. "
        f"(신규 {created_count}개 / 수정 {updated_count}개)",
    )
    return _redirect_back(request, "admin_students")


@admin_required
@require_POST
def admin_students_excel_upload(request):
    """Excel(.xlsx)로 수강생을 일괄 등록/수정한다. 이메일을 기준으로 기존 계정을 갱신한다."""
    uploaded = request.FILES.get("excel_file")
    if not uploaded:
        messages.error(request, "업로드할 Excel 파일을 선택해주세요.")
        return _redirect_back(request, "admin_students")

    if not uploaded.name.lower().endswith(".xlsx"):
        messages.error(request, ".xlsx 형식의 Excel 파일만 업로드할 수 있습니다.")
        return _redirect_back(request, "admin_students")

    if uploaded.size > 5 * 1024 * 1024:
        messages.error(request, "Excel 파일은 5MB 이하만 업로드할 수 있습니다.")
        return _redirect_back(request, "admin_students")

    try:
        from openpyxl import load_workbook
        workbook = load_workbook(uploaded, read_only=True, data_only=True)
        worksheet = workbook.active
    except Exception:
        messages.error(request, "Excel 파일을 읽을 수 없습니다. 파일이 손상되지 않았는지 확인해주세요.")
        return _redirect_back(request, "admin_students")

    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        messages.error(request, "Excel 파일에 데이터가 없습니다.")
        return _redirect_back(request, "admin_students")

    aliases = {
        "name": {"이름", "성명", "name", "studentname"},
        "email": {"이메일", "메일", "email", "emailaddress"},
        "password": {"비밀번호", "임시비밀번호", "password", "pw"},
        "affiliation": {"소속", "전공", "affiliation", "major", "department", "학과"},
        "team": {"팀", "팀명", "소속팀", "team", "teamname"},
        "is_active": {"활성", "상태", "활성상태", "isactive", "active"},
    }
    normalized_aliases = {
        key: {_normalize_excel_header(alias) for alias in values}
        for key, values in aliases.items()
    }
    headers = [_normalize_excel_header(value) for value in raw_headers]
    columns = {}
    for key, candidates in normalized_aliases.items():
        for index, header in enumerate(headers):
            if header in candidates:
                columns[key] = index
                break

    if "name" not in columns or "email" not in columns:
        messages.error(request, "Excel 첫 행에 '이름'과 '이메일' 열이 반드시 있어야 합니다.")
        return _redirect_back(request, "admin_students")

    current_round = _current_round()
    team_map = {
        team.name.strip().lower(): team
        for team in _round_teams(current_round)
    } if current_round else {}

    created = 0
    updated = 0
    skipped = 0
    errors = []

    def cell(row, key):
        index = columns.get(key)
        if index is None or index >= len(row):
            return ""
        value = row[index]
        return "" if value is None else str(value).strip()

    for excel_row, row in enumerate(rows, start=2):
        if not any(value not in (None, "") for value in row):
            continue

        name = cell(row, "name")
        email = cell(row, "email").lower()
        password = cell(row, "password")
        affiliation = cell(row, "affiliation")
        team_name = cell(row, "team")
        active_raw = cell(row, "is_active")

        try:
            if not name or not email:
                raise ValueError("이름과 이메일은 필수입니다.")
            if "@" not in email:
                raise ValueError("이메일 형식이 올바르지 않습니다.")

            active = _excel_bool(active_raw, default=True) if "is_active" in columns else None
            team = None
            if "team" in columns and team_name:
                if not current_round:
                    raise ValueError("현재 평가 회차가 없어 팀을 배정할 수 없습니다.")
                team = team_map.get(team_name.lower())
                if not team:
                    raise ValueError(f"현재 회차에서 '{team_name}' 팀을 찾을 수 없습니다.")

            with transaction.atomic():
                user = User.objects.filter(Q(email__iexact=email) | Q(username__iexact=email)).first()
                if user:
                    student = getattr(user, "student_profile", None)
                    if student is None:
                        raise ValueError("같은 이메일의 관리자/일반 계정이 이미 존재합니다.")
                    user.first_name = name
                    user.last_name = ""
                    user.email = email
                    user.username = email
                    if active is not None:
                        user.is_active = active
                    if password:
                        if len(password) < 8:
                            raise ValueError("비밀번호는 8자 이상이어야 합니다.")
                        user.set_password(password)
                    user.save()
                    if "affiliation" in columns:
                        student.affiliation = affiliation
                    if active is not None:
                        student.is_active = active
                    student.save()
                    if "team" in columns:
                        _sync_student_team(student, team, current_round)
                    updated += 1
                else:
                    if not password:
                        raise ValueError("신규 수강생은 임시 비밀번호가 필요합니다.")
                    if len(password) < 8:
                        raise ValueError("비밀번호는 8자 이상이어야 합니다.")
                    new_active = True if active is None else active
                    user = User.objects.create_user(
                        username=email,
                        email=email,
                        password=password,
                        first_name=name,
                        is_active=new_active,
                    )
                    student = Student.objects.create(user=user, affiliation=affiliation, is_active=new_active)
                    _sync_student_common_skills(student)
                    if "team" in columns:
                        _sync_student_team(student, team, current_round)
                    created += 1
        except (ValueError, IntegrityError) as exc:
            skipped += 1
            errors.append(f"{excel_row}행: {exc}")
        except Exception as exc:
            skipped += 1
            errors.append(f"{excel_row}행: 처리 중 오류가 발생했습니다. ({exc})")

    workbook.close()
    messages.success(request, f"Excel 업로드 완료: 신규 {created}명, 수정 {updated}명, 제외 {skipped}명")
    if errors:
        preview = " / ".join(errors[:5])
        if len(errors) > 5:
            preview += f" / 외 {len(errors) - 5}건"
        messages.warning(request, preview)
    return _redirect_back(request, "admin_students")

@admin_required
def admin_students_excel_export(request):
    """현재 수강생 목록을 다시 업로드 가능한 .xlsx 형식으로 내려준다."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    current_round = _current_round()
    students = list(Student.objects.select_related("user").order_by("user__first_name", "user__username"))
    membership_map = {}
    if current_round and students:
        memberships = TeamMembership.objects.filter(
            student_id__in=[student.id for student in students],
            team__evaluation_round=current_round,
        ).select_related("team")
        membership_map = {membership.student_id: membership.team.name for membership in memberships}

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "수강생"
    headers = ["이름", "이메일", "임시비밀번호", "소속", "팀명", "활성상태"]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    for student in students:
        worksheet.append([
            student.name,
            student.email,
            "",  # 보안을 위해 기존 비밀번호는 절대 내보내지 않는다.
            student.affiliation,
            membership_map.get(student.id, ""),
            "활성" if student.is_active and student.user.is_active else "비활성",
        ])

    widths = {"A": 18, "B": 30, "C": 18, "D": 20, "E": 18, "F": 12}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="students.xlsx"'
    return response

@admin_required
@require_POST
@transaction.atomic
def admin_student_create(request):
    if request.method != "POST":
        return _redirect_back(request, "admin_students")

    current_round = _current_round()
    teams = _round_teams(current_round)
    form = StudentCreateForm(request.POST, teams=teams)

    if not form.is_valid():
        for errors in form.errors.values():
            for error in errors:
                messages.error(request, error)
        return _redirect_back(request, "admin_students")

    name = form.cleaned_data["name"].strip()
    email = form.cleaned_data["email"]
    password = form.cleaned_data["password"]
    is_active = form.cleaned_data["is_active"]

    # 이메일 없는 관리용 수강생도 등록할 수 있도록 내부 username을 자동 생성한다.
    username = email or f"student_{uuid.uuid4().hex[:16]}"
    user = User(
        username=username,
        email=email,
        first_name=name,
        is_active=is_active,
    )
    if password:
        user.set_password(password)
    else:
        user.set_unusable_password()
    user.save()
    student = Student.objects.create(
        user=user,
        affiliation=form.cleaned_data["affiliation"].strip(),
        is_active=is_active,
    )
    _sync_student_common_skills(student)
    _sync_student_team(student, form.cleaned_data["team_id"], current_round)

    messages.success(request, f"{student.name} 수강생을 등록했습니다.")
    return _redirect_back(request, "admin_students")

@admin_required
@require_POST
@transaction.atomic
def admin_student_update(request, student_id):
    if request.method != "POST":
        return _redirect_back(request, "admin_students")

    student = get_object_or_404(Student.objects.select_related("user"), pk=student_id)
    current_round = _current_round()
    teams = _round_teams(current_round)
    form = StudentUpdateForm(request.POST, student=student, teams=teams)

    if not form.is_valid():
        for errors in form.errors.values():
            for error in errors:
                messages.error(request, error)
        return _redirect_back(request, "admin_students")

    user = student.user
    submitted_email = form.cleaned_data["email"]
    is_active = form.cleaned_data["is_active"]

    user.first_name = form.cleaned_data["name"].strip()
    user.last_name = ""

    # 수정 화면에서 이메일을 비워 제출하면 기존 이메일/로그인 ID를 그대로 유지한다.
    # 이름만 바꾸고 싶을 때 이메일을 다시 입력할 필요가 없다.
    if submitted_email:
        user.email = submitted_email
        user.username = submitted_email

    user.is_active = is_active
    if form.cleaned_data["password"]:
        user.set_password(form.cleaned_data["password"])
    user.save()

    student.affiliation = form.cleaned_data["affiliation"].strip()
    student.is_active = is_active
    student.save()
    _sync_student_team(student, form.cleaned_data["team_id"], current_round)

    messages.success(request, f"{student.name} 수강생 정보를 수정했습니다.")
    return _redirect_back(request, "admin_students")

@admin_required
@require_POST
@transaction.atomic
def admin_student_toggle_active(request, student_id):
    if request.method != "POST":
        return _redirect_back(request, "admin_students")

    student = get_object_or_404(Student.objects.select_related("user"), pk=student_id)
    new_state = not (student.is_active and student.user.is_active)
    student.is_active = new_state
    student.save(update_fields=["is_active", "updated_at"])
    student.user.is_active = new_state
    student.user.save(update_fields=["is_active"])

    state_label = "활성화" if new_state else "비활성화"
    messages.success(request, f"{student.name} 수강생을 {state_label}했습니다.")
    return _redirect_back(request, "admin_students")

@admin_required
def admin_student_detail(request, student_id):
    student = get_object_or_404(Student.objects.select_related("user"), pk=student_id)
    current_round = _current_round()
    current_team = None
    if current_round:
        membership = (
            TeamMembership.objects.filter(student=student, team__evaluation_round=current_round)
            .select_related("team")
            .first()
        )
        current_team = membership.team if membership else None

    memberships = list(
        TeamMembership.objects.filter(student=student)
        .select_related("team", "team__evaluation_round")
        .order_by("-team__evaluation_round__start_at")
    )
    results = list(
        StudentResult.objects.filter(student=student)
        .select_related("evaluation_round")
        .order_by("-evaluation_round__start_at")
    )

    team_required = 0
    personal_required = 0
    team_submitted = 0
    personal_submitted = 0
    team_saved = 0
    personal_saved = 0
    if current_round:
        if current_team:
            team_required = Team.objects.filter(evaluation_round=current_round, is_active=True).exclude(pk=current_team.pk).count()
            personal_required = max(
                TeamMembership.objects.filter(team=current_team).exclude(student=student).count(), 0
            )
        team_qs = TeamEvaluation.objects.filter(evaluation_round=current_round, evaluator=student)
        personal_qs = PersonalEvaluation.objects.filter(evaluation_round=current_round, evaluator=student)
        team_submitted = team_qs.filter(is_submitted=True).count()
        personal_submitted = personal_qs.filter(is_submitted=True).count()
        team_saved = team_qs.filter(is_submitted=False).count()
        personal_saved = personal_qs.filter(is_submitted=False).count()

    required_total = team_required + personal_required
    submitted_total = team_submitted + personal_submitted
    completion_percent = round((submitted_total / required_total) * 100) if required_total else 0

    social_accounts = []
    try:
        for account in student.user.socialaccount_set.all():
            provider_name = {"google": "Google", "kakao": "Kakao"}.get(account.provider, account.provider.title())
            social_accounts.append({"provider": account.provider, "name": provider_name})
    except Exception:
        social_accounts = []

    activities = []
    for ev in TeamEvaluation.objects.filter(evaluator=student).select_related("target_team", "evaluation_round").order_by("-updated_at")[:8]:
        activities.append({
            "at": ev.updated_at,
            "icon": "bi-people",
            "title": f"{ev.target_team.name} 팀 평가",
            "detail": f"{ev.evaluation_round.name} · {'제출 완료' if ev.is_submitted else '임시 저장'}",
        })
    for ev in PersonalEvaluation.objects.filter(evaluator=student).select_related("target_student__user", "evaluation_round").order_by("-updated_at")[:8]:
        activities.append({
            "at": ev.updated_at,
            "icon": "bi-person-check",
            "title": f"{ev.target_student.name} 개인 평가",
            "detail": f"{ev.evaluation_round.name} · {'제출 완료' if ev.is_submitted else '임시 저장'}",
        })
    activities = sorted(activities, key=lambda item: item["at"], reverse=True)[:8]

    admin_comments = list(
        AdminStudentComment.objects.filter(student=student)
        .select_related("evaluation_round", "created_by")
        .order_by("-evaluation_round__start_at", "-updated_at")
    )
    comment_rounds = list(EvaluationRound.objects.order_by("-start_at"))
    self_reviews = list(
        SelfProjectReview.objects.filter(student=student)
        .select_related("evaluation_round")
        .order_by("-evaluation_round__start_at")
    )
    skill_profiles = list(
        StudentSkill.objects.filter(student=student)
        .select_related("skill")
        .order_by("-score", "skill__name")
    )

    return render(request, "admin_ui/student_detail.html", _base_context(
        student=student,
        current_round=current_round,
        current_team=current_team,
        memberships=memberships,
        results=results,
        social_accounts=social_accounts,
        activities=activities,
        admin_comments=admin_comments,
        comment_rounds=comment_rounds,
        self_reviews=self_reviews,
        skill_profiles=skill_profiles,
        eval_stats={
            "team_required": team_required,
            "team_submitted": team_submitted,
            "team_saved": team_saved,
            "personal_required": personal_required,
            "personal_submitted": personal_submitted,
            "personal_saved": personal_saved,
            "required_total": required_total,
            "submitted_total": submitted_total,
            "completion_percent": completion_percent,
        },
    ))

@admin_required
@require_POST
@transaction.atomic
def admin_student_skill_save(request, student_id):
    """수강생 역량 프로필 항목을 생성하거나 점수/메모를 갱신한다."""
    student = get_object_or_404(Student.objects.select_related("user"), pk=student_id)
    skill_name = (request.POST.get("skill_name") or "").strip()
    raw_score = (request.POST.get("score") or "").strip()
    note = (request.POST.get("note") or "").strip()

    if not skill_name:
        messages.error(request, "역량 이름을 입력해주세요.")
        return redirect("admin_student_detail", student_id=student.id)
    if len(skill_name) > 80:
        messages.error(request, "역량 이름은 80자 이하로 입력해주세요.")
        return redirect("admin_student_detail", student_id=student.id)

    try:
        score = int(raw_score)
    except (TypeError, ValueError):
        score = -1
    if not 0 <= score <= 100:
        messages.error(request, "역량 점수는 0~100 사이로 입력해주세요.")
        return redirect("admin_student_detail", student_id=student.id)

    if len(note) > 300:
        messages.error(request, "역량 메모는 300자 이하로 입력해주세요.")
        return redirect("admin_student_detail", student_id=student.id)

    skill, skill_created = Skill.objects.get_or_create(name=skill_name)
    if skill_created:
        _sync_skill_to_students(skill)
    profile, created = StudentSkill.objects.update_or_create(
        student=student,
        skill=skill,
        defaults={"score": score, "note": note},
    )
    action = "추가" if created else "수정"
    messages.success(request, f"{skill.name} 역량을 {score}점으로 {action}했습니다.")
    return redirect("admin_student_detail", student_id=student.id)


@admin_required
@require_POST
@transaction.atomic
def admin_student_skill_delete(request, student_id, profile_id):
    student = get_object_or_404(Student, pk=student_id)
    profile = get_object_or_404(
        StudentSkill.objects.select_related("skill"),
        pk=profile_id,
        student=student,
    )
    skill_name = profile.skill.name
    profile.delete()
    messages.success(request, f"{skill_name} 역량을 프로필에서 제거했습니다.")
    return redirect("admin_student_detail", student_id=student.id)


@admin_required
@require_POST
@transaction.atomic
def admin_student_comment_save(request, student_id):
    student = get_object_or_404(Student.objects.select_related("user"), pk=student_id)
    round_id = (request.POST.get("evaluation_round_id") or "").strip()
    comment = (request.POST.get("comment") or "").strip()

    if not round_id:
        messages.error(request, "피드백을 남길 평가 회차를 선택해주세요.")
        return redirect("admin_student_detail", student_id=student.id)
    evaluation_round = get_object_or_404(EvaluationRound, pk=round_id)
    if not comment:
        messages.error(request, "학생에게 전달할 피드백 내용을 입력해주세요.")
        return redirect("admin_student_detail", student_id=student.id)
    if len(comment) > 2000:
        messages.error(request, "관리자 피드백은 2,000자 이하로 입력해주세요.")
        return redirect("admin_student_detail", student_id=student.id)

    feedback, created = AdminStudentComment.objects.update_or_create(
        evaluation_round=evaluation_round,
        student=student,
        defaults={
            "comment": comment,
            "created_by": request.user,
            "read_at": None,
        },
    )
    action = "등록" if created else "수정"
    messages.success(request, f"{student.name} 학생의 {evaluation_round.name} 피드백을 {action}했습니다.")
    return redirect("admin_student_detail", student_id=student.id)


@admin_required
@require_POST
@transaction.atomic
def admin_student_comment_delete(request, student_id, comment_id):
    student = get_object_or_404(Student, pk=student_id)
    feedback = get_object_or_404(AdminStudentComment, pk=comment_id, student=student)
    round_name = feedback.evaluation_round.name
    feedback.delete()
    messages.success(request, f"{student.name} 학생의 {round_name} 관리자 피드백을 삭제했습니다.")
    return redirect("admin_student_detail", student_id=student.id)


@admin_required
@require_POST
@transaction.atomic
def admin_student_reset_password(request, student_id):
    student = get_object_or_404(Student.objects.select_related("user"), pk=student_id)
    new_password = (request.POST.get("new_password") or "").strip()
    generated = False
    if not new_password:
        alphabet = string.ascii_letters + string.digits
        new_password = (
            secrets.choice(string.ascii_uppercase)
            + secrets.choice(string.ascii_lowercase)
            + secrets.choice(string.digits)
            + "".join(secrets.choice(alphabet) for _ in range(9))
        )
        generated = True
    if len(new_password) < 8:
        messages.error(request, "임시 비밀번호는 8자 이상이어야 합니다.")
        return _redirect_back(request, "admin_students")

    student.user.set_password(new_password)
    student.user.save(update_fields=["password"])
    _invalidate_user_sessions(student.user_id)

    if generated:
        messages.success(request, f"{student.name}의 비밀번호를 초기화했습니다. 임시 비밀번호: {new_password}")
    else:
        messages.success(request, f"{student.name}의 비밀번호를 입력한 값으로 초기화했습니다.")
    return _redirect_back(request, "admin_students")

@admin_required
@require_POST
@transaction.atomic
def admin_student_delete(request, student_id):
    if request.method != "POST":
        return _redirect_back(request, "admin_students")

    student = get_object_or_404(Student.objects.select_related("user"), pk=student_id)
    name = student.name
    user = student.user
    user.delete()
    messages.success(request, f"{name} 수강생을 삭제했습니다.")
    return _redirect_back(request, "admin_students")


@admin_required
@require_POST
@transaction.atomic
def admin_students_bulk_message_send(request):
    """학생 관리 화면에서 선택한 여러 학생에게 같은 시스템 내부 메시지를 보낸다."""
    raw_ids = request.POST.getlist("student_ids")
    student_ids = []
    for raw_id in raw_ids:
        try:
            student_ids.append(int(raw_id))
        except (TypeError, ValueError):
            continue

    # 중복 선택값 제거
    student_ids = list(dict.fromkeys(student_ids))
    if not student_ids:
        messages.error(request, "메시지를 받을 학생을 한 명 이상 선택해주세요.")
        return _redirect_back(request, "admin_students")

    students = list(
        Student.objects.select_related("user")
        .filter(id__in=student_ids)
        .order_by("user__first_name", "user__username")
    )
    if not students:
        messages.error(request, "선택한 학생 정보를 찾을 수 없습니다.")
        return _redirect_back(request, "admin_students")

    title = (request.POST.get("title") or "").strip()
    body = (request.POST.get("body") or "").strip()
    priority = (request.POST.get("priority") or InternalMessage.Priority.NORMAL).strip()

    if not title or not body:
        messages.error(request, "메시지 제목과 내용을 모두 입력해주세요.")
        return _redirect_back(request, "admin_students")
    if len(title) > 160:
        messages.error(request, "메시지 제목은 160자 이하로 입력해주세요.")
        return _redirect_back(request, "admin_students")
    if len(body) > 5000:
        messages.error(request, "메시지 내용은 5,000자 이하로 입력해주세요.")
        return _redirect_back(request, "admin_students")
    if priority not in InternalMessage.Priority.values:
        priority = Announcement.Priority.NORMAL

    InternalMessage.objects.bulk_create([
        InternalMessage(
            recipient=student,
            sender=request.user,
            title=title,
            body=body,
            priority=priority,
        )
        for student in students
    ])

    messages.success(request, f"선택한 {len(students)}명에게 시스템 메시지를 보냈습니다.")
    return _redirect_back(request, "admin_students")


@admin_required
@require_POST
@transaction.atomic
def admin_student_message_send(request, student_id):
    """학생 관리 화면에서 특정 학생에게 시스템 내부 메시지를 보낸다."""
    student = get_object_or_404(Student.objects.select_related("user"), pk=student_id)
    title = (request.POST.get("title") or "").strip()
    body = (request.POST.get("body") or "").strip()
    priority = (request.POST.get("priority") or InternalMessage.Priority.NORMAL).strip()

    if not title or not body:
        messages.error(request, "메시지 제목과 내용을 모두 입력해주세요.")
        return _redirect_back(request, "admin_students")
    if len(title) > 160:
        messages.error(request, "메시지 제목은 160자 이하로 입력해주세요.")
        return _redirect_back(request, "admin_students")
    if len(body) > 5000:
        messages.error(request, "메시지 내용은 5,000자 이하로 입력해주세요.")
        return _redirect_back(request, "admin_students")
    if priority not in InternalMessage.Priority.values:
        priority = Announcement.Priority.NORMAL

    InternalMessage.objects.create(
        recipient=student,
        sender=request.user,
        title=title,
        body=body,
        priority=priority,
    )
    messages.success(request, f"{student.name} 학생에게 시스템 메시지를 보냈습니다.")
    return _redirect_back(request, "admin_students")
