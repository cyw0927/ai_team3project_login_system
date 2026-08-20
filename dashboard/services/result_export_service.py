"""Excel export builder for evaluation results."""

from io import BytesIO

from django.db.models import Count
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..models import (
    StudentResult,
    Team,
    TeamEvaluation,
    TeamMembership,
    TeamResult,
)


def build_result_workbook(evaluation_round):
    """Return XLSX bytes for one evaluation round."""
    memberships = {
        membership.student_id: membership.team
        for membership in TeamMembership.objects.filter(
            team__evaluation_round=evaluation_round,
        ).select_related("team")
    }

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "종합 결과"
    headers = [
        "순위", "수강생", "이메일", "팀", "팀 점수", "개인 점수",
        "가중 합산", "관리자 보정", "보정 사유", "최종 점수", "집계 상태",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="E8EEF8")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    results = (
        StudentResult.objects.filter(evaluation_round=evaluation_round)
        .select_related("student__user")
        .order_by(
            "is_excluded",
            "rank",
            "student__user__first_name",
            "student__user__username",
        )
    )
    for result in results:
        team = memberships.get(result.student_id)
        sheet.append([
            result.rank if not result.is_excluded else "제외",
            result.student.name,
            result.student.user.email,
            team.name if team else "-",
            float(result.team_score),
            float(result.personal_score),
            float(result.base_score),
            float(result.adjustment_score),
            result.adjustment_reason,
            float(result.final_score),
            "집계 제외" if result.is_excluded else "정상",
        ])

    team_sheet = workbook.create_sheet("팀 결과")
    team_sheet.append(["순위", "팀", "인원", "제출 평가 수", "평균 점수", "집계 상태"])
    for cell in team_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    member_counts = dict(
        TeamMembership.objects.filter(team__evaluation_round=evaluation_round)
        .values("team_id")
        .annotate(c=Count("id"))
        .values_list("team_id", "c")
    )
    evaluation_counts = dict(
        TeamEvaluation.objects.filter(
            evaluation_round=evaluation_round,
            is_submitted=True,
        )
        .values("target_team_id")
        .annotate(c=Count("id"))
        .values_list("target_team_id", "c")
    )
    result_map = {
        result.team_id: result
        for result in TeamResult.objects.filter(evaluation_round=evaluation_round)
    }

    for team in Team.objects.filter(
        evaluation_round=evaluation_round,
        is_active=True,
    ).order_by("name"):
        result = result_map.get(team.id)
        team_sheet.append([
            result.rank if result and not result.is_excluded else (
                "제외" if result and result.is_excluded else "-"
            ),
            team.name,
            member_counts.get(team.id, 0),
            evaluation_counts.get(team.id, 0),
            float(result.score) if result else None,
            "집계 제외" if result and result.is_excluded else "정상",
        ])

    for current_sheet in (sheet, team_sheet):
        for column in current_sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column)
            current_sheet.column_dimensions[column[0].column_letter].width = min(
                max(max_len + 3, 12),
                32,
            )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
